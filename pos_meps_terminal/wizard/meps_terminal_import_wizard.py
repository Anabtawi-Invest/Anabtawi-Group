# -*- coding: utf-8 -*-
"""Import MEPS/ApexECR terminal lists (Tid/Mid/SecureKey per branch) from an Excel file.

Expected columns (header row, any order): TID, MID, Merchant Name, Branch, Secure Key.
This matches the sheet acquirers hand out per customer (e.g. "Anabtawi ECR.xlsx").
Every customer gets a different file - nothing here is customer-specific.
"""
import base64
import io
import logging

from odoo import fields, models
from odoo.exceptions import UserError
from odoo.tools.translate import _

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

_logger = logging.getLogger(__name__)

REQUIRED_HEADERS = ("TID", "MID", "Secure Key")
KNOWN_HEADERS = ("TID", "MID", "Merchant Name", "Branch", "Secure Key")


class MepsTerminalImportWizard(models.TransientModel):
    _name = "meps.terminal.import.wizard"
    _description = "Import MEPS Terminals from Excel"

    file = fields.Binary(string="MEPS Terminals File (.xlsx)", required=True)
    filename = fields.Char(string="Filename")
    auto_link_pos_config = fields.Boolean(
        string="Auto-link to POS Config by Branch Name",
        default=True,
        help="When exactly one POS config name contains the branch/merchant name, "
        "attach the payment method to it automatically. Ambiguous or missing "
        "matches are left for manual linking.",
    )
    result_message = fields.Text(string="Result", readonly=True)

    def _match_pos_config(self, label):
        PosConfig = self.env["pos.config"].sudo()
        key = (label or "").upper().strip()
        if not key:
            return PosConfig
        configs = PosConfig.search([])
        matches = configs.filtered(lambda c: key in (c.name or "").upper())
        return matches if len(matches) == 1 else PosConfig

    def action_import(self):
        self.ensure_one()
        if load_workbook is None:
            raise UserError(_("The 'openpyxl' Python library is required to import Excel files."))
        if not self.file:
            raise UserError(_("Please select a file to import."))

        try:
            workbook = load_workbook(io.BytesIO(base64.b64decode(self.file)), data_only=True)
        except Exception as exc:
            raise UserError(_("Could not read the Excel file: %s") % exc)

        rows = list(workbook.worksheets[0].iter_rows(values_only=True))
        if not rows:
            raise UserError(_("The file is empty."))

        header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        col_index = {name: header.index(name) for name in KNOWN_HEADERS if name in header}
        missing = [name for name in REQUIRED_HEADERS if name not in col_index]
        if missing:
            raise UserError(
                _("Missing required column(s): %(missing)s. Expected headers: %(all)s")
                % {"missing": ", ".join(missing), "all": ", ".join(KNOWN_HEADERS)}
            )

        PaymentMethod = self.env["pos.payment.method"].sudo()
        journal = PaymentMethod._meps_get_or_create_journal()

        created = updated = linked = unlinked = skipped = 0
        for row in rows[1:]:
            if not any(row):
                continue

            def cell(name):
                idx = col_index.get(name)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""

            tid, mid, secure_key = cell("TID"), cell("MID"), cell("Secure Key")
            merchant_name, branch = cell("Merchant Name"), cell("Branch")
            label = merchant_name or branch or tid

            if not (tid and mid and secure_key):
                skipped += 1
                continue

            existing = PaymentMethod.search([("meps_tid", "=", tid)], limit=1)
            if existing:
                existing.write({
                    "payment_method_type": "terminal",
                    "use_payment_terminal": "meps",
                    "meps_mid": mid,
                    "meps_secure_key": secure_key,
                })
                method = existing
                updated += 1
            else:
                vals = {
                    "name": "MEPS - %s" % label,
                    "payment_method_type": "terminal",
                    "use_payment_terminal": "meps",
                    "meps_tid": tid,
                    "meps_mid": mid,
                    "meps_secure_key": secure_key,
                    "meps_currency_code": "400",
                }
                if journal:
                    vals["journal_id"] = journal.id
                try:
                    method = PaymentMethod.create(vals)
                    created += 1
                except Exception:
                    _logger.exception("MEPS import: could not create payment method for Tid %s.", tid)
                    skipped += 1
                    continue

            if self.auto_link_pos_config:
                config = self._match_pos_config(branch or merchant_name)
                if config and method.id not in config.payment_method_ids.ids:
                    config.write({"payment_method_ids": [(4, method.id)]})
                    linked += 1
                elif not config:
                    unlinked += 1

        self.result_message = _(
            "Import done: %(created)s created, %(updated)s updated, %(linked)s auto-linked to a "
            "POS config, %(unlinked)s need manual linking, %(skipped)s skipped (missing/invalid data)."
        ) % {
            "created": created,
            "updated": updated,
            "linked": linked,
            "unlinked": unlinked,
            "skipped": skipped,
        }
        return {
            "type": "ir.actions.act_window",
            "res_model": "meps.terminal.import.wizard",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
