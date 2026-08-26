import base64
import io
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


class HrPayslipRunImportWizard(models.TransientModel):
    _name = 'hr.payslip.run.import.wizard'
    _description = 'Payrun Bulk Salary Adjustments Wizard'

    payrun_id = fields.Many2one('hr.payslip.run', string="Payrun", required=False, ondelete='cascade')
    excel_file = fields.Binary(string="Excel File", help="Upload the completed salary adjustment Excel file.")
    file_name = fields.Char(string="File Name")

    def _get_emp_code(self, emp):
        """Returns the primary employee code (employee_number)."""
        if not emp:
            return ""
        code = (
            getattr(emp, 'employee_number', False)
            or getattr(emp, 'sb_employee_number', False)
            or getattr(emp, 'registration_number', False)
            or getattr(emp, 'barcode', False)
            or getattr(emp, 'identification_id', False)
            or ''
        )
        return str(code).strip()

    def _get_target_payslips(self):
        """Helper to return relevant payslips from Payrun or Active Selection."""
        if self.payrun_id and self.payrun_id.slip_ids:
            return self.payrun_id.slip_ids
        active_ids = self._context.get('active_ids', [])
        if active_ids and self._context.get('active_model') == 'hr.payslip':
            return self.env['hr.payslip'].browse(active_ids)
        if active_ids and self._context.get('active_model') == 'hr.employee':
            employees = self.env['hr.employee'].browse(active_ids)
            return self.env['hr.payslip'].search([
                ('employee_id', 'in', employees.ids),
                ('state', 'in', ['draft', 'verify']),
            ])
        return self.env['hr.payslip']

    def _get_slip_actual_salary(self, slip):
        """Extracts the Actual Salary from the Salary Computation tab lines safely."""
        if not slip:
            return 0.0

        if "line_ids" in slip._fields and slip.line_ids:
            actual_line = slip.line_ids.filtered(
                lambda l: (l.name and l.name.strip().lower() == 'actual salary')
                or (l.code and l.code.upper() in ['ACTUAL', 'ACTUAL_SALARY', 'ACTUAL_SAL'])
            )
            if actual_line:
                return actual_line[0].total
            net_line = slip.line_ids.filtered(
                lambda l: l.code and l.code.upper() in ['NET', 'GROSS', 'BASIC']
            )
            if net_line:
                return net_line[0].total

        if "wage" in slip._fields and getattr(slip, 'wage', False):
            return slip.wage

        if "version_id" in slip._fields and getattr(slip, 'version_id', False):
            version = slip.version_id
            if getattr(version, 'contract_wage', False):
                return version.contract_wage

        if "contract_id" in slip._fields and getattr(slip, 'contract_id', False):
            contract = slip.contract_id
            if getattr(contract, 'wage', False):
                return contract.wage

        if getattr(slip, 'employee_id', False):
            emp = slip.employee_id
            if "wage" in emp._fields and getattr(emp, 'wage', False):
                return emp.wage
            if "contract_id" in emp._fields and getattr(emp, 'contract_id', False):
                contract = emp.contract_id
                if getattr(contract, 'wage', False):
                    return contract.wage

        return 0.0

    def action_export_template(self):
        """Generates and downloads the Excel template with Employee Details, Actual Salary, Adjustment Type, and Amount."""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("The 'openpyxl' Python package is required to generate Excel files."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Adjustments"

        headers = [
            "Employee Code", 
            "Employee Name", 
            "Actual Salary", 
            "Adjustment Type", 
            "Adjustment Amount", 
            "Notes"
        ]

        header_fill = openpyxl.styles.PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        slips = self._get_target_payslips().sorted(key=lambda s: s.employee_id.name or '')
        if not slips:
            active_emp_ids = self._context.get('active_ids', [])
            if active_emp_ids and self._context.get('active_model') == 'hr.employee':
                employees = self.env['hr.employee'].browse(active_emp_ids)
            else:
                employees = self.env['hr.employee'].search([('active', '=', True)])

            for emp in employees.sorted(key=lambda e: e.name or ''):
                emp_code = self._get_emp_code(emp)
                wage = 0.0
                if "wage" in emp._fields and emp.wage:
                    wage = emp.wage
                elif "contract_id" in emp._fields and emp.contract_id and getattr(emp.contract_id, 'wage', False):
                    wage = emp.contract_id.wage
                ws.append([
                    emp_code, 
                    emp.name, 
                    round(wage, 3), 
                    "Partial Salary Payment", 
                    0.0, 
                    ""
                ])
        else:
            for slip in slips:
                emp = slip.employee_id
                emp_code = self._get_emp_code(emp)
                actual_salary = self._get_slip_actual_salary(slip)
                ws.append([
                    emp_code, 
                    emp.name, 
                    round(actual_salary, 3), 
                    "Partial Salary Payment", 
                    0.0, 
                    ""
                ])

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 26
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 30

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=3).number_format = '#,##0.000'
            ws.cell(row=row, column=5).number_format = '#,##0.000'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_data = base64.b64encode(output.read())

        payrun_name = self.payrun_id.name if self.payrun_id else 'Batch'
        filename = f"Salary_Adjustments_{payrun_name}.xlsx".replace(" ", "_")
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': file_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _resolve_attachment_type(self, target_model_obj, field_name, type_str):
        """Dynamically matches the salary adjustment type from the Excel sheet against hr.salary.attachment.type."""
        field_obj = target_model_obj._fields.get(field_name)
        if not field_obj or field_obj.type != 'many2one':
            return False

        comodel_name = field_obj.comodel_name
        if not comodel_name or comodel_name not in self.env:
            return False

        CoModel = self.env[comodel_name].sudo()
        c_fields = CoModel._fields

        # 1. If user provided a Type Name or Code in the Excel file, search for it
        if type_str:
            clean_type = str(type_str).strip()
            domain = []
            if 'code' in c_fields and 'name' in c_fields:
                domain = ['|', ('code', '=ilike', clean_type), ('name', '=ilike', clean_type)]
            elif 'code' in c_fields:
                domain = [('code', '=ilike', clean_type)]
            elif 'name' in c_fields:
                domain = [('name', '=ilike', clean_type)]

            found = CoModel.search(domain, limit=1) if domain else False
            if not found and ('name' in c_fields or 'code' in c_fields):
                # Fuzzy partial matching
                domain_fuzzy = []
                if 'name' in c_fields and 'code' in c_fields:
                    domain_fuzzy = ['|', ('code', 'ilike', clean_type), ('name', 'ilike', clean_type)]
                elif 'name' in c_fields:
                    domain_fuzzy = [('name', 'ilike', clean_type)]
                elif 'code' in c_fields:
                    domain_fuzzy = [('code', 'ilike', clean_type)]
                found = CoModel.search(domain_fuzzy, limit=1)
            if found:
                return found.id

        # 2. Fallback to standard codes in system
        fallback_codes = ['PA_pay', 'PARTIAL_PAYMENT', 'SAL_PAY1', 'com_lon', 'LOAN', 'ADVANCE', 'OTHER']
        found = False
        if 'code' in c_fields:
            found = CoModel.search([('code', 'in', fallback_codes)], limit=1)
        if not found and 'name' in c_fields:
            found = CoModel.search([('name', 'ilike', 'partial')], limit=1)
        if not found:
            found = CoModel.search([], limit=1)

        # 3. Create default if none exists
        if not found and 'name' in c_fields:
            try:
                vals = {'name': 'Partial Salary Payment'}
                if 'code' in c_fields:
                    vals['code'] = 'PA_pay'
                found = CoModel.create(vals)
            except Exception as e:
                _logger.warning("Could not auto-create type on %s: %s", comodel_name, str(e))
                found = False

        return found.id if found else False

    def _create_employee_salary_adjustment(self, emp, type_str, amount, note_text):
        """Creates or updates a Salary Adjustment / Attachment record under the Employee Profile."""
        today = fields.Date.today()
        created_or_updated = False
        errors = []
        clean_note = (note_text or type_str or _("Salary partial payment")).strip()

        # 1. Standard Odoo Salary Attachment model (hr.salary.attachment)
        if 'hr.salary.attachment' in self.env:
            try:
                with self.env.cr.savepoint():
                    Attachment = self.env['hr.salary.attachment'].sudo()
                    fields_dict = Attachment._fields

                    domain = []
                    if 'employee_ids' in fields_dict:
                        domain.append(('employee_ids', 'in', [emp.id]))
                    elif 'employee_id' in fields_dict:
                        domain.append(('employee_id', '=', emp.id))

                    if 'description' in fields_dict:
                        domain.append(('description', '=', clean_note))
                    elif 'name' in fields_dict:
                        domain.append(('name', '=', clean_note))

                    existing = Attachment.search(domain, limit=1) if domain else False

                    vals = {}
                    # Many2many / Many2one employee assignment
                    if 'employee_ids' in fields_dict:
                        vals['employee_ids'] = [(6, 0, [emp.id])]
                    elif 'employee_id' in fields_dict and not fields_dict['employee_id'].readonly:
                        vals['employee_id'] = emp.id

                    if 'description' in fields_dict:
                        vals['description'] = clean_note
                    if 'name' in fields_dict:
                        vals['name'] = clean_note

                    if 'monthly_amount' in fields_dict:
                        vals['monthly_amount'] = amount
                    if 'total_amount' in fields_dict:
                        vals['total_amount'] = amount
                    if 'amount' in fields_dict:
                        vals['amount'] = amount

                    if 'date_start' in fields_dict:
                        vals['date_start'] = today
                    elif 'date' in fields_dict:
                        vals['date'] = today

                    if 'company_id' in fields_dict:
                        vals['company_id'] = emp.company_id.id if emp.company_id else self.env.company.id

                    # Resolve type field (deduction_type_id / attachment_type_id / type_id)
                    for type_fname in ['deduction_type_id', 'attachment_type_id', 'payslip_input_type_id', 'input_type_id', 'type_id']:
                        if type_fname in fields_dict:
                            type_val = self._resolve_attachment_type(Attachment, type_fname, type_str)
                            if type_val:
                                vals[type_fname] = type_val
                                break

                    if existing:
                        existing.write(vals)
                    else:
                        Attachment.create(vals)
                    created_or_updated = True

            except Exception as e:
                err_msg = f"hr.salary.attachment failed for {emp.name}: {str(e)}"
                _logger.warning(err_msg)
                errors.append(err_msg)

        # 2. Check custom Salary Adjustment models if present
        for model_name in ['hr.salary.adjustment', 'sb.hr.salary.adjustment', 'hr.employee.salary.adjustment']:
            if model_name in self.env:
                try:
                    with self.env.cr.savepoint():
                        AdjModel = self.env[model_name].sudo()
                        adj_fields = AdjModel._fields
                        vals = {}
                        if 'employee_id' in adj_fields:
                            vals['employee_id'] = emp.id
                        elif 'employee_ids' in adj_fields:
                            vals['employee_ids'] = [(6, 0, [emp.id])]

                        if 'monthly_amount' in adj_fields:
                            vals['monthly_amount'] = amount
                        if 'total_amount' in adj_fields:
                            vals['total_amount'] = amount
                        if 'amount' in adj_fields:
                            vals['amount'] = amount
                        elif 'payslip_amount' in adj_fields:
                            vals['payslip_amount'] = amount

                        if 'note' in adj_fields:
                            vals['note'] = clean_note
                        if 'description' in adj_fields:
                            vals['description'] = clean_note
                        if 'name' in adj_fields:
                            vals['name'] = clean_note

                        if 'date_start' in adj_fields:
                            vals['date_start'] = today
                        elif 'date' in adj_fields:
                            vals['date'] = today

                        if 'company_id' in adj_fields:
                            vals['company_id'] = emp.company_id.id if emp.company_id else self.env.company.id

                        for type_fname in ['type_id', 'input_type_id', 'adjustment_type_id', 'deduction_type_id']:
                            if type_fname in adj_fields:
                                type_val = self._resolve_attachment_type(AdjModel, type_fname, type_str)
                                if type_val:
                                    vals[type_fname] = type_val
                                    break

                        AdjModel.create(vals)
                        created_or_updated = True

                except Exception as e:
                    err_msg = f"{model_name} failed for {emp.name}: {str(e)}"
                    _logger.warning(err_msg)
                    errors.append(err_msg)

        return created_or_updated, errors

    def action_import_adjustments(self):
        """Reads the uploaded Excel file and creates Salary Adjustments on employee profiles."""
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("The 'openpyxl' Python package is required to read Excel files."))
        if not self.excel_file:
            raise UserError(_("Please select an Excel file to upload."))

        file_content = base64.b64decode(self.excel_file)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise UserError(_("Failed to read Excel file: %s") % str(e))

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            raise UserError(_("The uploaded Excel file contains no data rows."))

        header_row = [str(cell).strip() if cell is not None else '' for cell in rows[0]]

        code_col_idx = None
        name_col_idx = None
        type_col_idx = None
        amount_col_idx = None
        notes_col_idx = None

        for idx, col_name in enumerate(header_row):
            c_lower = str(col_name).lower().strip()
            # Employee Code
            if any(k in c_lower for k in ['code', 'badge', 'number', 'emp id', 'employee id', 'الرقم', 'رقم']):
                if code_col_idx is None:
                    code_col_idx = idx
            # Employee Name
            elif any(k in c_lower for k in ['name', 'employee name', 'اسم', 'الاسم']):
                if name_col_idx is None:
                    name_col_idx = idx
            # Adjustment Type
            elif any(k in c_lower for k in ['adjustment type', 'type', 'salary adjustment type', 'نوع', 'نوع الحركة', 'نوع التعديل', 'نوع الخصم']):
                type_col_idx = idx
            # Notes / Description
            elif any(k in c_lower for k in ['note', 'notes', 'description', 'ملاحظات', 'بيان']):
                notes_col_idx = idx
            # Amount
            elif any(k in c_lower for k in ['adjustment amount', 'partial payment', 'amount', 'partial', 'loan', 'advance', 'payment', 'مبلغ', 'قيمة', 'دفعة', 'سلفة']):
                if 'actual' not in c_lower and 'wage' not in c_lower and 'salary' not in c_lower:
                    amount_col_idx = idx
                elif amount_col_idx is None:
                    amount_col_idx = idx

        # Default fallbacks based on column positions
        if code_col_idx is None:
            code_col_idx = 0
        if name_col_idx is None:
            name_col_idx = 1 if len(header_row) > 1 else None

        # 6-column format vs 5-column format
        if len(header_row) >= 6:
            if type_col_idx is None:
                type_col_idx = 3
            if amount_col_idx is None:
                amount_col_idx = 4
            if notes_col_idx is None:
                notes_col_idx = 5
        else:
            if amount_col_idx is None:
                amount_col_idx = 3 if len(header_row) > 3 else 2
            if notes_col_idx is None and len(header_row) > 4:
                notes_col_idx = 4

        # Build comprehensive index of all employees
        all_employees = self.env['hr.employee'].sudo().search([])
        emp_by_code = {}
        for emp in all_employees:
            for field_name in ['employee_number', 'sb_employee_number', 'registration_number', 'barcode', 'identification_id', 'pin']:
                val = getattr(emp, field_name, False)
                if val:
                    val_str = str(val).strip()
                    emp_by_code[val_str] = emp
                    emp_by_code[val_str.lstrip('0')] = emp
            emp_by_code[str(emp.id)] = emp
            if emp.name:
                emp_by_code[emp.name.strip().lower()] = emp

        missing_codes = []
        all_errors = []
        updated_count = 0
        total_amount = 0.0

        for row_idx, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue

            raw_code = row[code_col_idx] if code_col_idx < len(row) else None
            raw_name = row[name_col_idx] if (name_col_idx is not None and name_col_idx < len(row)) else None

            emp = None
            emp_code_str = ""

            if raw_code is not None:
                emp_code_str = str(raw_code).strip()
                if emp_code_str.endswith('.0'):
                    emp_code_str = emp_code_str[:-2]
                emp = emp_by_code.get(emp_code_str) or emp_by_code.get(emp_code_str.lstrip('0'))

            if not emp and raw_name:
                clean_name = str(raw_name).strip().lower()
                emp = emp_by_code.get(clean_name)

            if not emp:
                display_label = emp_code_str or str(raw_name or '') or f"Row {row_idx}"
                missing_codes.append(f"Row {row_idx}: '{display_label}'")
                continue

            # Parse Type, Amount, and Notes
            raw_type = row[type_col_idx] if (type_col_idx is not None and type_col_idx < len(row)) else ""
            type_str = str(raw_type).strip() if raw_type is not None else ""

            raw_amount = row[amount_col_idx] if (amount_col_idx is not None and amount_col_idx < len(row)) else 0.0
            try:
                if isinstance(raw_amount, (int, float)):
                    adj_amount = float(raw_amount)
                elif isinstance(raw_amount, str):
                    clean_str = raw_amount.replace(',', '').replace('JOD', '').replace('JD', '').replace('$', '').strip()
                    adj_amount = float(clean_str) if clean_str else 0.0
                else:
                    adj_amount = 0.0
            except (ValueError, TypeError):
                adj_amount = 0.0

            raw_notes = row[notes_col_idx] if (notes_col_idx is not None and notes_col_idx < len(row)) else ""
            note_str = str(raw_notes).strip() if raw_notes is not None else ""

            if adj_amount > 0.0:
                success, errors = self._create_employee_salary_adjustment(emp, type_str, adj_amount, note_str)
                if success:
                    updated_count += 1
                    total_amount += adj_amount
                if errors:
                    all_errors.extend(errors)

        # Log to payrun chatter if linked
        payrun = self.payrun_id
        if payrun:
            log_msg = _(
                "<b>Salary Adjustments Imported</b><br/>"
                "• File: <code>%s</code><br/>"
                "• Employees Updated: <b>%d</b><br/>"
                "• Total Amount: <b>%.3f JOD</b>"
            ) % (self.file_name or 'Uploaded Excel', updated_count, total_amount)
            if missing_codes:
                log_msg += "<br/><br/><b style='color:red;'>Unmatched Rows (%d):</b><br/>%s" % (
                    len(missing_codes), "<br/>".join(missing_codes[:10])
                )
            if all_errors:
                log_msg += "<br/><br/><b style='color:red;'>Errors (%d):</b><br/>%s" % (
                    len(all_errors), "<br/>".join(all_errors[:5])
                )
            payrun.message_post(body=log_msg)

        if updated_count == 0 and (missing_codes or all_errors):
            err_details = []
            if missing_codes:
                err_details.append(_("Unmatched Employees (%d): %s") % (len(missing_codes), ", ".join(missing_codes[:5])))
            if all_errors:
                err_details.append(_("Errors: %s") % "; ".join(all_errors[:3]))
            raise UserError(_("No adjustments were created.\n\n%s") % "\n".join(err_details))

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _("Salary Adjustments Created"),
                'message': _(
                    "Successfully created Salary Adjustments for %d employees (Total: %.3f JOD)."
                ) % (updated_count, total_amount),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.client', 'tag': 'reload'},
            },
        }
