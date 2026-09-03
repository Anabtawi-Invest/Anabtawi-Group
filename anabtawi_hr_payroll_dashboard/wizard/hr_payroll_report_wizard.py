# -*- coding: utf-8 -*-
import base64
import io
import logging
from datetime import datetime, date, timedelta

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class HrPayrollReportWizard(models.TransientModel):
    _name = "hr.payroll.report.wizard"
    _description = "HR & Payroll Unified Excel Report Generator"

    date_from = fields.Date(string="Date From", required=True, default=lambda self: fields.Date.today().replace(day=1))
    date_to = fields.Date(string="Date To", required=True, default=lambda self: (fields.Date.today().replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1))
    payrun_id = fields.Many2one("hr.payslip.run", string="Payrun Batch", help="Filter by specific payroll batch")
    company_id = fields.Many2one("res.company", string="Company", default=lambda self: self.env.company, required=True)
    department_ids = fields.Many2many("hr.department", string="Departments")
    report_type = fields.Selection([
        ("all", "Complete HR & Payroll Package (All 4 Sheets)"),
        ("payroll", "1. Monthly Payroll List"),
        ("overtime", "2. Monthly Overtime Analysis"),
        ("bank", "3. Bank Transfer File"),
        ("attendance", "4. Attendance Reconciliation"),
    ], string="Report Scope", default="all", required=True)

    company_source_account = fields.Char(string="Company Bank Account Number", default="0126046919500", help="Used in Bank Transfer file (BT/LBT)")

    excel_file = fields.Binary(string="Excel File", readonly=True)
    file_name = fields.Char(string="File Name")

    def _get_emp_code(self, emp):
        if not emp:
            return ""
        return str(
            getattr(emp, "registration_number", False)
            or getattr(emp, "employee_number", False)
            or getattr(emp, "barcode", False)
            or getattr(emp, "identification_id", False)
            or emp.id
        ).strip()

    def _get_emp_name_en(self, emp):
        if not emp:
            return ""
        return str(getattr(emp, "legal_name", False) or emp.name or "").strip()

    def action_export_xlsx(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("The openpyxl Python library is required for generating Excel reports."))

        slip_domain = [
            ("state", "not in", ["cancel"]),
            ("company_id", "=", self.company_id.id),
        ]
        if self.payrun_id:
            slip_domain.append(("payslip_run_id", "=", self.payrun_id.id))
        else:
            slip_domain.extend([
                ("date_from", "<=", self.date_to),
                ("date_to", ">=", self.date_from),
            ])

        if self.department_ids:
            slip_domain.append(("department_id", "in", self.department_ids.ids))

        payslips = self.env["hr.payslip"].search(slip_domain, order="department_id, employee_id")
        if not payslips:
            raise UserError(_("No payslip records found for the selected criteria."))

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        font_header_title = Font(name="Arial", size=14, bold=True, color="1E3A8A")
        font_sub_title = Font(name="Arial", size=10, italic=True, color="475569")
        font_col_header = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        font_sub_header = Font(name="Arial", size=9, bold=True, color="1E293B")
        font_data = Font(name="Arial", size=9)
        font_total = Font(name="Arial", size=9, bold=True, color="000000")

        fill_primary_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        fill_alw_header = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
        fill_ded_header = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
        fill_sub_header = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        fill_grand_total = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")

        thin_side = Side(border_style="thin", color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        double_bottom = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(border_style="double", color="000000"))

        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        period_title = self.payrun_id.name if self.payrun_id else f"{self.date_from.strftime('%m/%Y')}"

        # -------------------------------------------------------------
        # SHEET 1: Monthly Payroll List
        # -------------------------------------------------------------
        if self.report_type in ["all", "payroll"]:
            ws_pay = wb.create_sheet(title="Monthly Payroll List")
            ws_pay.views.sheetView[0].showGridLines = True

            ws_pay.cell(row=1, column=16, value=f"Monthly Payroll List {period_title}").font = font_header_title
            ws_pay.cell(row=2, column=16, value="مياومة,عقد غير محدد المدة,عقد وزارة العمل,متدرب,عقد محدد المدة").font = font_sub_title
            ws_pay.cell(row=3, column=16, value="Final Copy-Current Employees").font = font_sub_title

            headers_group = [
                (1, 10, "Employee & Salary Information", fill_primary_header),
                (11, 17, "Allowances (العلاوات والبدلات)", fill_alw_header),
                (18, 18, "Gross Salary", fill_primary_header),
                (19, 32, "Deductions & Taxes (الخصومات والاقتطاعات)", fill_ded_header),
                (33, 33, "Net Salary", fill_primary_header),
            ]
            for start_c, end_c, title, fill in headers_group:
                if start_c != end_c:
                    ws_pay.merge_cells(start_row=5, start_column=start_c, end_row=5, end_column=end_c)
                c_cell = ws_pay.cell(row=5, column=start_c, value=title)
                c_cell.font = font_col_header
                c_cell.fill = fill
                c_cell.alignment = align_center

            cols = [
                "Property", "دائرة", "قسم", "فرع", "Emp No.", "Emp Name", 
                "Working Days", "Basic Salary", "Actual basic salary", "Due Salary",
                "علاوة تكليف", "مكافأة مالية-شهرية", "عمل اضافي", "بدل تنقلات", "عائد ضمان", "علاوات أخرى", "Total Allowances",
                "Gross Salary",
                "دفعة أولى", "دفعة ثانية", "فرق راتب", "قرض شركة", "City Ledger", "برنامج وطني", "تأمين صحي", "عقوبات وغرامات",
                "ضمان اجتماعي", "ضريبة الدخل", "مخالفات وخصومات", "خصم ساعات", "سلف أخرى", "Total Deductions",
                "Net Salary"
            ]
            for idx, col_name in enumerate(cols, start=1):
                cell = ws_pay.cell(row=6, column=idx, value=col_name)
                cell.font = font_sub_header
                cell.fill = fill_sub_header
                cell.alignment = align_center
                cell.border = border_all

            row_idx = 7
            for slip in payslips:
                emp = slip.employee_id
                dept_name = (slip.department_id or emp.department_id).name or _("Unassigned")
                branch_name = getattr(emp, "work_location_id", False) and emp.work_location_id.name or dept_name
                division_name = getattr(emp.department_id, "parent_id", False) and emp.department_id.parent_id.name or dept_name

                basic_w = getattr(slip, "basic_wage", 0.0) or getattr(slip, "wage", 0.0) or (slip.contract_id and slip.contract_id.wage) or (slip.version_id and getattr(slip.version_id, "contract_wage", 0.0)) or 0.0
                actual_w = basic_w
                due_w = basic_w

                alw_takleef = 0.0
                alw_bonus = 0.0
                alw_ot = 0.0
                alw_trans = 0.0
                alw_ss = 0.0
                alw_other = 0.0

                ded_pay1 = 0.0
                ded_pay2 = 0.0
                ded_diff = 0.0
                ded_loan = 0.0
                ded_city = 0.0
                ded_nat = 0.0
                ded_ins = 0.0
                ded_pen = 0.0
                ded_ss = 0.0
                ded_tax = 0.0
                ded_viol = 0.0
                ded_hours = 0.0
                ded_other = 0.0

                gross_val = getattr(slip, "gross_wage", 0.0) or basic_w
                net_val = getattr(slip, "net_wage", 0.0) or basic_w

                if hasattr(slip, "line_ids") and slip.line_ids:
                    for l in slip.line_ids:
                        c_code = (l.code or "").upper().strip()
                        c_name = (l.name or "").lower()
                        amt = l.total or 0.0

                        if c_code == "NET": net_val = amt
                        elif c_code in ["GROSS", "GRS"]: gross_val = amt
                        elif "تكليف" in c_name: alw_takleef += amt
                        elif "مكافأة" in c_name or "مكافاه" in c_name: alw_bonus += amt
                        elif "إضافي" in c_name or "اضافي" in c_name or "OT" in c_code: alw_ot += amt
                        elif "تنقل" in c_name or "مواصلات" in c_name: alw_trans += amt
                        elif "عائد ضمان" in c_name: alw_ss += amt
                        elif (l.category_id and l.category_id.code in ["ALW", "ALLOWANCE"]) and c_code not in ["NET", "GROSS", "BASIC"]:
                            alw_other += amt
                        
                        elif "دفعة اولى" in c_name or "دفعة 1" in c_name: ded_pay1 += abs(amt)
                        elif "دفعة ثانية" in c_name or "دفعة 2" in c_name: ded_pay2 += abs(amt)
                        elif "فرق راتب" in c_name: ded_diff += abs(amt)
                        elif "قرض" in c_name or "سلفة" in c_name or "LOAN" in c_code: ded_loan += abs(amt)
                        elif "city ledger" in c_name: ded_city += abs(amt)
                        elif "وطني" in c_name: ded_nat += abs(amt)
                        elif "تأمين" in c_name or "تامين" in c_name: ded_ins += abs(amt)
                        elif "عقوبات" in c_name or "غرامات" in c_name: ded_pen += abs(amt)
                        elif "ضمان" in c_name or "SS" in c_code or "SSC" in c_code: ded_ss += abs(amt)
                        elif "ضريبة" in c_name or "TAX" in c_code: ded_tax += abs(amt)
                        elif "مخالفات" in c_name: ded_viol += abs(amt)
                        elif "خصم ساعات" in c_name or "تأخير" in c_name or "LATE" in c_code: ded_hours += abs(amt)
                        elif (l.category_id and l.category_id.code in ["DED", "DEDUCTION"]) and c_code not in ["NET", "GROSS"]:
                            ded_other += abs(amt)

                tot_alw = alw_takleef + alw_bonus + alw_ot + alw_trans + alw_ss + alw_other
                tot_ded = ded_pay1 + ded_pay2 + ded_diff + ded_loan + ded_city + ded_nat + ded_ins + ded_pen + ded_ss + ded_tax + ded_viol + ded_hours + ded_other
                if not gross_val: gross_val = basic_w + tot_alw
                if not net_val: net_val = gross_val - tot_ded

                w_days = 30.0
                if hasattr(slip, "worked_days_line_ids") and slip.worked_days_line_ids:
                    w_days = sum(wd.number_of_days for wd in slip.worked_days_line_ids if wd.code != "OUT") or 30.0

                row_values = [
                    self.company_id.name,
                    division_name,
                    dept_name,
                    branch_name,
                    self._get_emp_code(emp),
                    emp.name,
                    round(w_days, 2),
                    round(basic_w, 3),
                    round(actual_w, 3),
                    round(due_w, 3),
                    round(alw_takleef, 3),
                    round(alw_bonus, 3),
                    round(alw_ot, 3),
                    round(alw_trans, 3),
                    round(alw_ss, 3),
                    round(alw_other, 3),
                    round(tot_alw, 3),
                    round(gross_val, 3),
                    round(ded_pay1, 3),
                    round(ded_pay2, 3),
                    round(ded_diff, 3),
                    round(ded_loan, 3),
                    round(ded_city, 3),
                    round(ded_nat, 3),
                    round(ded_ins, 3),
                    round(ded_pen, 3),
                    round(ded_ss, 3),
                    round(ded_tax, 3),
                    round(ded_viol, 3),
                    round(ded_hours, 3),
                    round(ded_other, 3),
                    round(tot_ded, 3),
                    round(net_val, 3),
                ]

                for col_i, val in enumerate(row_values, start=1):
                    c = ws_pay.cell(row=row_idx, column=col_i, value=val)
                    c.font = font_data
                    c.border = border_all
                    if isinstance(val, (int, float)):
                        c.alignment = align_right
                        c.number_format = "#,##0.000" if col_i not in [7] else "0.00"
                    else:
                        c.alignment = align_left

                row_idx += 1

            ws_pay.cell(row=row_idx, column=5, value="Grand Total").font = font_total
            ws_pay.cell(row=row_idx, column=5).fill = fill_grand_total
            for col_i in range(7, 34):
                col_letter = get_column_letter(col_i)
                sum_cell = ws_pay.cell(row=row_idx, column=col_i, value=f"=SUM({col_letter}7:{col_letter}{row_idx-1})")
                sum_cell.font = font_total
                sum_cell.fill = fill_grand_total
                sum_cell.border = double_bottom
                sum_cell.alignment = align_right
                sum_cell.number_format = "#,##0.000" if col_i != 7 else "0.00"

            for col in ws_pay.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws_pay.column_dimensions[col_letter].width = max(max_len + 3, 12)

        # -------------------------------------------------------------
        # SHEET 2: Monthly Overtime Analysis
        # -------------------------------------------------------------
        if self.report_type in ["all", "overtime"]:
            ws_ot = wb.create_sheet(title="Monthly Overtime")
            ws_ot.views.sheetView[0].showGridLines = True

            ws_ot.cell(row=1, column=1, value=f"Monthly Overtime {period_title}").font = font_header_title
            ws_ot.cell(row=2, column=1, value="Final Copy").font = font_sub_title

            ot_headers = [
                "Emp No.", "Emp Name", "Department / Branch", "Document Ref", "OT Date",
                "Overtime Rate", "Pay Rate / Hr", "Salary", "Accrued Hours", "Paid Hours", "Days", "Amount"
            ]
            for c_i, h_name in enumerate(ot_headers, start=1):
                cell = ws_ot.cell(row=5, column=c_i, value=h_name)
                cell.font = font_col_header
                cell.fill = fill_primary_header
                cell.alignment = align_center
                cell.border = border_all

            ot_row = 6
            for slip in payslips:
                emp = slip.employee_id
                dept_name = (slip.department_id or emp.department_id).name or ""
                basic_w = getattr(slip, "basic_wage", 0.0) or getattr(slip, "wage", 0.0) or 0.0
                pay_rate = round(basic_w / 240.0, 3) if basic_w else 0.0

                ot_hours = 0.0
                ot_amt = 0.0
                if hasattr(slip, "worked_days_line_ids") and slip.worked_days_line_ids:
                    ot_hours = sum(wd.number_of_hours for wd in slip.worked_days_line_ids if "OT" in (wd.code or "").upper() or "EXTRA" in (wd.code or "").upper() or "اضافي" in (wd.name or ""))
                
                if hasattr(slip, "line_ids") and slip.line_ids:
                    ot_amt = sum(l.total for l in slip.line_ids if "OT" in (l.code or "").upper() or "OVERTIME" in (l.code or "").upper() or "اضافي" in (l.name or ""))
                
                if not ot_amt and ot_hours:
                    ot_amt = round(ot_hours * pay_rate * 1.25, 3)

                ot_vals = [
                    self._get_emp_code(emp),
                    emp.name,
                    dept_name,
                    slip.number or slip.name or "OT-RUN",
                    str(slip.date_to or self.date_to),
                    1.25,
                    pay_rate,
                    round(basic_w, 3),
                    round(ot_hours, 2),
                    round(ot_hours, 2),
                    round(ot_hours / 8.0, 2),
                    round(ot_amt, 3),
                ]

                for col_i, val in enumerate(ot_vals, start=1):
                    c = ws_ot.cell(row=ot_row, column=col_i, value=val)
                    c.font = font_data
                    c.border = border_all
                    if isinstance(val, (int, float)):
                        c.alignment = align_right
                        c.number_format = "#,##0.000" if col_i in [7, 8, 12] else "0.00"
                    else:
                        c.alignment = align_left
                ot_row += 1

            ws_ot.cell(row=ot_row, column=2, value="Total Overtime").font = font_total
            ws_ot.cell(row=ot_row, column=2).fill = fill_grand_total
            for col_i in [9, 10, 11, 12]:
                c_let = get_column_letter(col_i)
                sum_cell = ws_ot.cell(row=ot_row, column=col_i, value=f"=SUM({c_let}6:{c_let}{ot_row-1})")
                sum_cell.font = font_total
                sum_cell.fill = fill_grand_total
                sum_cell.border = double_bottom
                sum_cell.alignment = align_right
                sum_cell.number_format = "#,##0.000" if col_i == 12 else "0.00"

            for col in ws_ot.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                c_let = get_column_letter(col[0].column)
                ws_ot.column_dimensions[c_let].width = max(max_len + 3, 12)

        # -------------------------------------------------------------
        # SHEET 3: Bank Transfer File (Exact Bank Format)
        # -------------------------------------------------------------
        if self.report_type in ["all", "bank"]:
            ws_bank = wb.create_sheet(title="Bank File Factory")
            ws_bank.views.sheetView[0].showGridLines = True

            bank_row = 1
            src_acc = self.company_source_account or "0126046919500"
            val_date = (self.date_to + timedelta(days=5)).strftime("%d/%m/%Y")
            desc_text = f"SALARY PAYMENT {self.date_to.strftime('%m %Y')}"

            for slip in payslips:
                emp = slip.employee_id
                net_amt = getattr(slip, "net_wage", 0.0) or 0.0
                if hasattr(slip, "line_ids") and slip.line_ids:
                    net_line = slip.line_ids.filtered(lambda l: l.code == "NET")
                    if net_line:
                        net_amt = net_line[0].total

                bank_acc = emp.bank_account_id or (getattr(emp, "bank_account_ids", False) and emp.bank_account_ids[0])
                acc_num = bank_acc.acc_number if bank_acc else ""
                bic = bank_acc.bank_id.bic if bank_acc and bank_acc.bank_id else "ARABJOAXXXX"
                if not bic: bic = "ARABJOAXXXX"

                txn_type = "LBT" if (acc_num and acc_num.strip().upper().startswith("JO")) else "BT"
                emp_name_en = self._get_emp_name_en(emp) or emp.name

                b_vals = [
                    txn_type,
                    src_acc,
                    acc_num or "CASH_NO_IBAN",
                    emp_name_en.upper(),
                    bic.upper(),
                    "13",
                    val_date,
                    "JOD",
                    round(net_amt, 3),
                    "",
                    "",
                    desc_text,
                    self._get_emp_code(emp),
                ]

                for col_i, val in enumerate(b_vals, start=1):
                    c = ws_bank.cell(row=bank_row, column=col_i, value=val)
                    c.font = font_data
                    if isinstance(val, (int, float)):
                        c.alignment = align_right
                        c.number_format = "0.000"
                    else:
                        c.alignment = align_left

                bank_row += 1

            for col in ws_bank.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                c_let = get_column_letter(col[0].column)
                ws_bank.column_dimensions[c_let].width = max(max_len + 2, 14)

        # -------------------------------------------------------------
        # SHEET 4: Attendance Reconciliation
        # -------------------------------------------------------------
        if self.report_type in ["all", "attendance"]:
            ws_att = wb.create_sheet(title="Attendance Reconciliation")
            ws_att.views.sheetView[0].showGridLines = True

            ws_att.cell(row=1, column=1, value="Attendance Reconciliation").font = font_header_title
            ws_att.cell(row=2, column=1, value=f"Period: {self.date_from} to {self.date_to}").font = font_sub_title

            att_headers = [
                "Number", "Name", "Department", "Scheduled Hours", "Approved Hours",
                "Extra Hours", "Subtractive Hours", "Reconciliation Hours"
            ]
            for c_i, h_name in enumerate(att_headers, start=1):
                cell = ws_att.cell(row=4, column=c_i, value=h_name)
                cell.font = font_col_header
                cell.fill = fill_primary_header
                cell.alignment = align_center
                cell.border = border_all

            att_row = 5
            for slip in payslips:
                emp = slip.employee_id
                dept_name = (slip.department_id or emp.department_id).name or ""

                w_days = 30.0
                ot_hrs = 0.0
                late_hrs = 0.0

                if hasattr(slip, "worked_days_line_ids") and slip.worked_days_line_ids:
                    w_days = sum(wd.number_of_days for wd in slip.worked_days_line_ids if wd.code != "OUT") or 30.0
                    ot_hrs = sum(wd.number_of_hours for wd in slip.worked_days_line_ids if "OT" in (wd.code or "").upper() or "EXTRA" in (wd.code or "").upper() or "اضافي" in (wd.name or ""))
                    late_hrs = sum(wd.number_of_hours for wd in slip.worked_days_line_ids if "LATE" in (wd.code or "").upper() or "تأخير" in (wd.name or "") or "خصم" in (wd.name or ""))

                sched_hrs = round(w_days * 8.0, 2)
                approved_hrs = round(sched_hrs + ot_hrs - late_hrs, 2)
                recon_hrs = round(ot_hrs - late_hrs, 2)

                def _fmt_hrs(val):
                    sign = "-" if val < 0 else ""
                    val = abs(val)
                    h = int(val)
                    m = int(round((val - h) * 60))
                    return f"{sign}{h:02d}:{m:02d}"

                att_vals = [
                    self._get_emp_code(emp),
                    emp.name,
                    dept_name,
                    _fmt_hrs(sched_hrs),
                    _fmt_hrs(approved_hrs),
                    _fmt_hrs(ot_hrs),
                    _fmt_hrs(late_hrs),
                    _fmt_hrs(recon_hrs),
                ]

                for col_i, val in enumerate(att_vals, start=1):
                    c = ws_att.cell(row=att_row, column=col_i, value=val)
                    c.font = font_data
                    c.border = border_all
                    c.alignment = align_center if col_i in [1, 4, 5, 6, 7, 8] else align_left

                att_row += 1

            for col in ws_att.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                c_let = get_column_letter(col[0].column)
                ws_att.column_dimensions[c_let].width = max(max_len + 3, 14)

        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        file_data = base64.b64encode(fp.read())
        fp.close()

        file_name = f"HR_Payroll_Unified_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        self.write({
            "excel_file": file_data,
            "file_name": file_name,
        })

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/?model=hr.payroll.report.wizard&id={self.id}&field=excel_file&download=true&filename={file_name}",
            "target": "self",
        }
