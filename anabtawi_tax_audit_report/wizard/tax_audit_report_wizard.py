import io
import datetime
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class TaxAuditReportWizard(models.TransientModel):
    _name = "tax.audit.report.wizard"
    _description = "Tax Audit & Declaration Report Wizard"

    @api.model
    def _default_date_from(self):
        today = fields.Date.today()
        return today.replace(day=1)

    @api.model
    def _default_date_to(self):
        today = fields.Date.today()
        next_month = today.replace(day=28) + datetime.timedelta(days=4)
        return next_month - datetime.timedelta(days=next_month.day)

    date_from = fields.Date(string="Start Date", required=True, default=_default_date_from)
    date_to = fields.Date(string="End Date", required=True, default=_default_date_to)
    company_id = fields.Many2one(
        "res.company",
        string="Company",
        required=True,
        default=lambda self: self.env.company,
    )
    currency_id = fields.Many2one(
        "res.currency",
        string="Currency",
        required=True,
        default=lambda self: self.env.company.currency_id,
    )
    target_state = fields.Selection(
        [
            ("posted", "Posted Invoices & Bills Only"),
            ("all", "All Statuses (Draft & Posted)"),
        ],
        string="Invoice State",
        required=True,
        default="posted",
    )
    tax_jurisdiction = fields.Char(
        string="Tax Jurisdiction",
        required=True,
        default="Jordan (16% Standard Rate)",
    )
    tax_rate = fields.Float(
        string="Default Tax Rate (%)",
        required=True,
        default=16.0,
    )

    excel_file = fields.Binary(string="Excel Output File", readonly=True)
    filename = fields.Char(string="Excel Filename", readonly=True)

    def _get_report_data(self):
        """Fetches transactions and categorizes into 4 streams:
        1. Local Sales Invoice (out_invoice non-POS)
        2. POS Sales Invoice (out_invoice POS / pos.order)
        3. Trade Purchase Bill (in_invoice COGS / Raw Materials)
        4. Expense Invoice / Bill (in_invoice Expenses)
        """
        self.ensure_one()
        domain = [
            ("company_id", "=", self.company_id.id),
            ("move_type", "in", ["out_invoice", "out_refund", "in_invoice", "in_refund"]),
        ]

        if self.target_state == "posted":
            domain.append(("state", "=", "posted"))
        else:
            domain.append(("state", "!=", "cancel"))

        if self.date_from:
            domain.append(("invoice_date", ">=", self.date_from))
        if self.date_to:
            domain.append(("invoice_date", "<=", self.date_to))

        moves = self.env["account.move"].search(domain, order="invoice_date asc, name asc")

        itemized_data = []
        summary_counts = {
            "local_sales": {"count": 0, "base": 0.0, "tax": 0.0},
            "pos_sales": {"count": 0, "base": 0.0, "tax": 0.0},
            "trade_purchases": {"count": 0, "base": 0.0, "tax": 0.0},
            "expense_bills": {"count": 0, "base": 0.0, "tax": 0.0},
        }

        for move in moves:
            is_pos = False
            if hasattr(move, "pos_order_ids") and move.pos_order_ids:
                is_pos = True
            elif hasattr(move, "journal_id") and "pos" in (move.journal_id.code or "").lower():
                is_pos = True

            # Determine category
            if move.move_type in ["out_invoice", "out_refund"]:
                sign = 1 if move.move_type == "out_invoice" else -1
                if is_pos:
                    cat_key = "pos_sales"
                    category_label = "POS Sales Invoice"
                    tax_name = f"POS Tax {int(self.tax_rate)}%"
                else:
                    cat_key = "local_sales"
                    category_label = "Local Sales Invoice"
                    tax_name = f"Sales Tax {int(self.tax_rate)}%"
            else:
                sign = 1 if move.move_type == "in_invoice" else -1
                # Check line account types for COGS vs Expense
                is_cogs = False
                for line in move.invoice_line_ids:
                    account_type = getattr(line.account_id, "account_type", "") or ""
                    if "direct_costs" in account_type or "expense_direct" in account_type or "cogs" in (line.account_id.code or "").lower():
                        is_cogs = True
                        break

                if is_cogs:
                    cat_key = "trade_purchases"
                    category_label = "Trade Purchase Bill"
                    tax_name = f"Purchase Tax {int(self.tax_rate)}%"
                else:
                    cat_key = "expense_bills"
                    category_label = "Expense Invoice / Bill"
                    tax_name = f"Expense Tax {int(self.tax_rate)}%"

            base = (move.amount_untaxed or 0.0) * sign
            tax = (move.amount_tax or 0.0) * sign
            total = (move.amount_total or 0.0) * sign
            rate_dec = self.tax_rate / 100.0

            summary_counts[cat_key]["count"] += 1
            summary_counts[cat_key]["base"] += base
            summary_counts[cat_key]["tax"] += tax

            itemized_data.append({
                "date": str(move.invoice_date or move.date or ""),
                "inv_number": move.name or _("Draft"),
                "partner": move.partner_id.display_name or _("Generic Partner"),
                "category": category_label,
                "cat_key": cat_key,
                "tax_name": tax_name,
                "tax_rate": rate_dec,
                "taxable_base": base,
                "tax_amount": tax,
                "total_amount": total,
                "status": move.state.capitalize() if move.state else "Draft",
            })

        return {
            "summary": summary_counts,
            "details": itemized_data,
        }

    def _generate_excel_workbook(self):
        if not openpyxl:
            raise UserError(_("The openpyxl library is required to generate Excel reports."))

        data = self._get_report_data()
        itemized_list = data["details"]

        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "Tax Declaration Summary"
        ws_details = wb.create_sheet(title="Itemized Invoices Audit")

        ws_summary.views.sheetView[0].showGridLines = True
        ws_details.views.sheetView[0].showGridLines = True

        NAVY_HEADER = "1F4E79"
        STEEL_BLUE = "2F5597"
        ICE_BLUE = "D9E1F2"
        LIGHT_ZEBRA = "F2F2F2"
        BORDER_GRAY = "D9D9D9"
        TOTAL_BG = "E6EEF8"

        font_title = Font(name="Calibri", size=16, bold=True, color="1F4E79")
        font_subtitle = Font(name="Calibri", size=10, italic=True, color="595959")
        font_th = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=11)

        fill_navy = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
        fill_steel = PatternFill(start_color=STEEL_BLUE, end_color=STEEL_BLUE, fill_type="solid")
        fill_ice = PatternFill(start_color=ICE_BLUE, end_color=ICE_BLUE, fill_type="solid")
        fill_zebra = PatternFill(start_color=LIGHT_ZEBRA, end_color=LIGHT_ZEBRA, fill_type="solid")
        fill_total = PatternFill(start_color=TOTAL_BG, end_color=TOTAL_BG, fill_type="solid")

        thin_border_side = Side(border_style="thin", color=BORDER_GRAY)
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        top_thin_bottom_double = Border(top=Side(border_style="thin", color="000000"), bottom=Side(border_style="double", color="000000"))

        # --- SHEET 1: SUMMARY ---
        ws_summary['A1'] = "COMPREHENSIVE TAX DECLARATION & AUDIT SUMMARY"
        ws_summary['A1'].font = font_title
        curr_code = self.currency_id.name or "JOD"
        as_of_str = self.date_to.strftime('%B %Y') if self.date_to else datetime.date.today().strftime('%B %Y')
        ws_summary['A2'] = f"Generated for Tax Reporting Period ({self.date_from} to {self.date_to}) | Currency: {curr_code} | As of: {as_of_str}"
        ws_summary['A2'].font = font_subtitle

        headers_summary = ["Tax Declaration Stream / Category", "Invoice Count", f"Taxable Base ({curr_code})", "Tax Rate", f"Tax Amount ({curr_code})"]
        for col_num, h in enumerate(headers_summary, 1):
            cell = ws_summary.cell(row=4, column=col_num)
            cell.value = h
            cell.font = font_th
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center" if col_num in [2,4] else "right" if col_num in [3,5] else "left", vertical="center")

        # Output Tax (Sales)
        ws_summary['A5'] = "A. OUTPUT TAX (SALES & REVENUE)"
        ws_summary.merge_cells('A5:E5')
        ws_summary['A5'].font = font_bold
        ws_summary['A5'].fill = fill_ice

        summary_rows_a = [
            ("Local Wholesale / Commercial Sales Invoices", "=COUNTIFS('Itemized Invoices Audit'!D:D, \"Local Sales Invoice\")", "=SUMIFS('Itemized Invoices Audit'!G:G, 'Itemized Invoices Audit'!D:D, \"Local Sales Invoice\")", f"{self.tax_rate:.1f}%", "=SUMIFS('Itemized Invoices Audit'!I:I, 'Itemized Invoices Audit'!D:D, \"Local Sales Invoice\")"),
            ("POS Retail Sales Invoices / Daily Z-Reports", "=COUNTIFS('Itemized Invoices Audit'!D:D, \"POS Sales Invoice\")", "=SUMIFS('Itemized Invoices Audit'!G:G, 'Itemized Invoices Audit'!D:D, \"POS Sales Invoice\")", f"{self.tax_rate:.1f}%", "=SUMIFS('Itemized Invoices Audit'!I:I, 'Itemized Invoices Audit'!D:D, \"POS Sales Invoice\")"),
        ]

        for idx, r_data in enumerate(summary_rows_a, start=6):
            for c_idx, val in enumerate(r_data, start=1):
                cell = ws_summary.cell(row=idx, column=c_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "#,##0"
                elif c_idx in [3, 5]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
                elif c_idx == 4:
                    cell.alignment = Alignment(horizontal="center")

        ws_summary['A8'] = "Subtotal: Output Tax (Sales)"
        ws_summary['B8'] = "=SUM(B6:B7)"
        ws_summary['C8'] = "=SUM(C6:C7)"
        ws_summary['D8'] = "-"
        ws_summary['E8'] = "=SUM(E6:E7)"
        for c in range(1, 6):
            cell = ws_summary.cell(row=8, column=c)
            cell.font = font_bold
            cell.fill = fill_total
            cell.border = thin_border
            if c in [3, 5]: cell.number_format = "#,##0.00"
            if c == 2: cell.number_format = "#,##0"

        # Input Tax (Purchases & Expenses)
        ws_summary['A10'] = "B. INPUT TAX (PURCHASES & EXPENSES)"
        ws_summary.merge_cells('A10:E10')
        ws_summary['A10'].font = font_bold
        ws_summary['A10'].fill = fill_ice

        summary_rows_b = [
            ("Trade Purchases / Raw Materials Vendor Bills", "=COUNTIFS('Itemized Invoices Audit'!D:D, \"Trade Purchase Bill\")", "=SUMIFS('Itemized Invoices Audit'!G:G, 'Itemized Invoices Audit'!D:D, \"Trade Purchase Bill\")", f"{self.tax_rate:.1f}%", "=SUMIFS('Itemized Invoices Audit'!I:I, 'Itemized Invoices Audit'!D:D, \"Trade Purchase Bill\")"),
            ("Operational Expense Invoices / Bills (Rent, Utilities, Services)", "=COUNTIFS('Itemized Invoices Audit'!D:D, \"Expense Invoice / Bill\")", "=SUMIFS('Itemized Invoices Audit'!G:G, 'Itemized Invoices Audit'!D:D, \"Expense Invoice / Bill\")", f"{self.tax_rate:.1f}%", "=SUMIFS('Itemized Invoices Audit'!I:I, 'Itemized Invoices Audit'!D:D, \"Expense Invoice / Bill\")"),
        ]

        for idx, r_data in enumerate(summary_rows_b, start=11):
            for c_idx, val in enumerate(r_data, start=1):
                cell = ws_summary.cell(row=idx, column=c_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                if c_idx == 2:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "#,##0"
                elif c_idx in [3, 5]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"
                elif c_idx == 4:
                    cell.alignment = Alignment(horizontal="center")

        ws_summary['A13'] = "Subtotal: Input Tax (Purchases & Expenses)"
        ws_summary['B13'] = "=SUM(B11:B12)"
        ws_summary['C13'] = "=SUM(C11:C12)"
        ws_summary['D13'] = "-"
        ws_summary['E13'] = "=SUM(E11:E12)"
        for c in range(1, 6):
            cell = ws_summary.cell(row=13, column=c)
            cell.font = font_bold
            cell.fill = fill_total
            cell.border = thin_border
            if c in [3, 5]: cell.number_format = "#,##0.00"
            if c == 2: cell.number_format = "#,##0"

        # Net Tax Settlement
        ws_summary['A15'] = "NET TAX PAYABLE / (REFUNDABLE) [A - B]"
        ws_summary['B15'] = "=B8+B13"
        ws_summary['C15'] = "=C8-C13"
        ws_summary['D15'] = "-"
        ws_summary['E15'] = "=E8-E13"
        for c in range(1, 6):
            cell = ws_summary.cell(row=15, column=c)
            cell.font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
            cell.fill = PatternFill(start_color="D9E1F2", fill_type="solid")
            cell.border = top_thin_bottom_double
            if c in [3, 5]: cell.number_format = "#,##0.00"
            if c == 2: cell.number_format = "#,##0"

        # --- SHEET 2: ITEMIZED DETAILS ---
        ws_details['A1'] = "DETAILED TAX INVOICES & BILLS AUDIT REGISTER"
        ws_details['A1'].font = font_title
        ws_details['A2'] = "Complete transaction breakdown for Local Sales, POS Orders, Raw Material Purchases, and Expense Invoices"
        ws_details['A2'].font = font_subtitle

        detail_headers = [
            "Date", "Invoice / Bill #", "Partner / Customer / Vendor", "Transaction Category", 
            "Tax Name", "Tax Rate", f"Taxable Base ({curr_code})", f"Tax Amount ({curr_code})", f"Total with Tax ({curr_code})", "Status"
        ]

        for col_num, h in enumerate(detail_headers, 1):
            cell = ws_details.cell(row=4, column=col_num)
            cell.value = h
            cell.font = font_th
            cell.fill = fill_steel
            cell.alignment = Alignment(horizontal="center" if col_num in [1, 2, 4, 5, 6, 10] else "right" if col_num in [7, 8, 9] else "left", vertical="center")

        start_row = 5
        for row_offset, item in enumerate(itemized_list):
            r_idx = start_row + row_offset
            row_vals = [
                item["date"],
                item["inv_number"],
                item["partner"],
                item["category"],
                item["tax_name"],
                item["tax_rate"],
                item["taxable_base"],
                f"=G{r_idx}*F{r_idx}",
                f"=G{r_idx}+H{r_idx}",
                item["status"],
            ]
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws_details.cell(row=r_idx, column=col_idx, value=val)
                cell.font = font_regular
                cell.border = thin_border
                if r_idx % 2 == 0:
                    cell.fill = fill_zebra
                
                if col_idx in [1, 2, 4, 5, 10]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx == 6:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "0.0%"
                elif col_idx in [7, 8, 9]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = "#,##0.00"

        end_row = start_row + len(itemized_list) - 1 if itemized_list else start_row
        tot_row = end_row + 1 if itemized_list else start_row + 1

        ws_details.cell(row=tot_row, column=1, value="GRAND TOTAL AUDIT")
        ws_details.merge_cells(start_row=tot_row, start_column=1, end_row=tot_row, end_column=6)
        ws_details.cell(row=tot_row, column=1).alignment = Alignment(horizontal="right")
        ws_details.cell(row=tot_row, column=1).font = font_bold

        sum_start = start_row
        sum_end = end_row if itemized_list else start_row
        ws_details.cell(row=tot_row, column=7, value=f"=SUM(G{sum_start}:G{sum_end})")
        ws_details.cell(row=tot_row, column=8, value=f"=SUM(H{sum_start}:H{sum_end})")
        ws_details.cell(row=tot_row, column=9, value=f"=SUM(I{sum_start}:I{sum_end})")

        for col_idx in range(1, 11):
            c = ws_details.cell(row=tot_row, column=col_idx)
            c.font = font_bold
            c.fill = fill_total
            c.border = top_thin_bottom_double
            if col_idx in [7, 8, 9]:
                c.number_format = "#,##0.00"

        for ws in [ws_summary, ws_details]:
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        ws_summary.column_dimensions['A'].width = 46
        ws_details.column_dimensions['C'].width = 38
        ws_details.column_dimensions['D'].width = 24

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def action_generate_excel(self):
        import base64
        excel_bytes = self._generate_excel_workbook()
        file_b64 = base64.b64encode(excel_bytes)
        filename = f"Tax_Audit_Report_{fields.Date.today()}.xlsx"
        
        self.write({
            'excel_file': file_b64,
            'filename': filename,
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/anabtawi_tax_audit/download_excel?wizard_id={self.id}',
            'target': 'self',
        }

    def action_generate_pdf(self):
        return self.env.ref("anabtawi_tax_audit_report.action_report_tax_audit_pdf").report_action(self)
