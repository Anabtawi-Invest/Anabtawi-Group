import base64
import io
from datetime import datetime, time

import pytz

from odoo import api, fields, models, _
from odoo.exceptions import AccessError, UserError, ValidationError

try:
    import openpyxl
except ImportError:
    openpyxl = None

STATEMENT_COLUMNS = {
    "order_id": "Order Id",
    "date": "Date / Time",
    "status": "Status",
    "total_sale": "Total Sale",
    "voucher_cost_to_talabat": "Voucher Cost To Talabat",
    "tpro_cost_covered": "tPro cost covered by Talabat for discounted orders",
    "sponsored_deal_fees": "Sponsored Deal Fees",
    "payment_handling_charges": "Payment Handling Charges",
    "payment_handling_charges_vat": "Payment Handling Charges VAT",
    "loyalty_charges": "Loyalty Charges",
    "commission": "Commission ",
    "commission_vat": "Commission VAT",
}
STATEMENT_CANCELLED_STATUS = "Charged Cancelled"


class OnlineCampaignSettlement(models.Model):
    _name = "online.campaign.settlement"
    _description = "Aggregator Campaign Settlement"
    _inherit = ["mail.thread", "mail.activity.mixin"]
    _order = "date_end desc, id desc"

    name = fields.Char(required=True, default="/", copy=False, tracking=True)
    state = fields.Selection(
        [("draft", "Draft"), ("confirmed", "Confirmed"), ("reconciled", "Reconciled")],
        default="draft", required=True, tracking=True, index=True,
    )
    aggregator_id = fields.Many2one(
        "online.campaign.aggregator", required=True, check_company=True, tracking=True, index=True
    )
    company_id = fields.Many2one(
        "res.company", required=True, default=lambda self: self.env.company, index=True
    )
    currency_id = fields.Many2one("res.currency", required=True)
    date_start = fields.Date(required=True, tracking=True)
    date_end = fields.Date(required=True, tracking=True)
    statement_reference = fields.Char(tracking=True)
    statement_file = fields.Binary(attachment=True)
    statement_filename = fields.Char()
    order_count = fields.Integer(readonly=True)
    line_count = fields.Integer(readonly=True)

    expected_customer_collections = fields.Monetary(currency_field="currency_id", readonly=True)
    expected_contribution = fields.Monetary(currency_field="currency_id", readonly=True)
    expected_commission = fields.Monetary(currency_field="currency_id", readonly=True)
    expected_net_settlement = fields.Monetary(
        currency_field="currency_id", compute="_compute_totals", store=True
    )

    actual_customer_collections = fields.Monetary(currency_field="currency_id", tracking=True)
    actual_contribution = fields.Monetary(currency_field="currency_id", tracking=True)
    actual_commission = fields.Monetary(currency_field="currency_id", tracking=True)
    adjustment_amount = fields.Monetary(
        currency_field="currency_id", tracking=True,
        help="Signed statement adjustment: positive increases the settlement, negative reduces it."
    )
    actual_tpro_loyalty_charges = fields.Monetary(
        string="Actual TPro (Loyalty Charges)", currency_field="currency_id", tracking=True,
        help="Sum of the statement's Loyalty Charges column.",
    )
    actual_tpro_cost_covered = fields.Monetary(
        string="Actual TPro Cost Covered by Talabat", currency_field="currency_id", tracking=True,
        help="Sum of the statement's tPro cost covered by Talabat for discounted orders column.",
    )
    actual_sponsored_deal_fees = fields.Monetary(
        currency_field="currency_id", tracking=True,
        help="Sum of the statement's Sponsored Deal Fees column.",
    )
    actual_payment_handling_charges = fields.Monetary(
        currency_field="currency_id", tracking=True,
        help="Sum of the statement's Payment Handling Charges column.",
    )
    actual_payment_handling_charges_vat = fields.Monetary(
        currency_field="currency_id", tracking=True,
        help="Sum of the statement's Payment Handling Charges VAT column.",
    )
    statement_order_count = fields.Integer(
        readonly=True, copy=False,
        help="Number of order rows read from the imported statement file.",
    )
    cancelled_charged_order_count = fields.Integer(
        string="Cancelled but Charged Orders", readonly=True, copy=False,
        help="Orders the statement marks as cancelled but still billed commission on.",
    )
    cancelled_charged_commission = fields.Monetary(
        string="Commission Paid on Cancelled Orders",
        currency_field="currency_id", readonly=True, copy=False,
        help="Commission (incl. VAT) charged on orders that were cancelled, "
        "already included in Actual Commission above. Shown separately as a "
        "pure loss: no sale was completed.",
    )
    actual_net_received = fields.Monetary(
        currency_field="currency_id", compute="_compute_totals", store=True
    )
    variance_amount = fields.Monetary(
        currency_field="currency_id", compute="_compute_totals", store=True
    )
    variance_percent = fields.Float(compute="_compute_totals", store=True, digits=(16, 4))
    customer_collections_mismatch = fields.Boolean(compute="_compute_totals", store=True)
    contribution_mismatch = fields.Boolean(compute="_compute_totals", store=True)
    commission_mismatch = fields.Boolean(compute="_compute_totals", store=True)
    variance_reason = fields.Text(tracking=True)
    bank_statement_line_id = fields.Many2one(
        "account.bank.statement.line", string="Bank Statement Line", check_company=True
    )
    account_move_id = fields.Many2one(
        "account.move", string="Settlement/Commission Entry", check_company=True
    )
    reconciled_by = fields.Many2one("res.users", readonly=True, copy=False)
    reconciled_at = fields.Datetime(readonly=True, copy=False)

    @api.onchange("aggregator_id")
    def _onchange_aggregator(self):
        if self.aggregator_id:
            self.company_id = self.aggregator_id.company_id
            self.currency_id = self.aggregator_id.company_id.currency_id

    @api.model_create_multi
    def create(self, vals_list):
        for values in vals_list:
            if values.get("aggregator_id"):
                aggregator = self.env["online.campaign.aggregator"].browse(values["aggregator_id"])
                values.setdefault("company_id", aggregator.company_id.id)
                values.setdefault("currency_id", aggregator.company_id.currency_id.id)
            if values.get("name", "/") == "/":
                values["name"] = _(
                    "%(aggregator)s %(start)s - %(end)s",
                    aggregator=self.env["online.campaign.aggregator"].browse(
                        values.get("aggregator_id")
                    ).display_name,
                    start=values.get("date_start", ""), end=values.get("date_end", ""),
                )
        return super().create(vals_list)

    def write(self, values):
        locked_fields = {
            "aggregator_id", "company_id", "currency_id", "date_start", "date_end",
            "actual_customer_collections", "actual_contribution", "actual_commission",
            "actual_tpro_loyalty_charges", "actual_tpro_cost_covered",
            "actual_sponsored_deal_fees", "actual_payment_handling_charges",
            "actual_payment_handling_charges_vat", "statement_order_count",
            "cancelled_charged_order_count", "cancelled_charged_commission",
            "adjustment_amount", "statement_reference", "statement_file",
            "bank_statement_line_id", "account_move_id", "variance_reason",
            "expected_customer_collections", "expected_contribution", "expected_commission",
            "order_count", "line_count", "state",
        }
        if locked_fields.intersection(values) and self.filtered(
            lambda settlement: settlement.state == "reconciled"
        ):
            raise UserError(_("A reconciled settlement is locked."))
        return super().write(values)

    @api.depends(
        "expected_customer_collections", "expected_contribution", "expected_commission",
        "actual_customer_collections", "actual_contribution", "actual_commission",
        "adjustment_amount",
    )
    def _compute_totals(self):
        for settlement in self:
            settlement.expected_net_settlement = (
                settlement.expected_customer_collections
                + settlement.expected_contribution
                - settlement.expected_commission
            )
            settlement.actual_net_received = (
                settlement.actual_customer_collections
                + settlement.actual_contribution
                - settlement.actual_commission
                + settlement.adjustment_amount
            )
            settlement.variance_amount = (
                settlement.actual_net_received - settlement.expected_net_settlement
            )
            settlement.variance_percent = (
                settlement.variance_amount / settlement.expected_net_settlement
                if settlement.expected_net_settlement else 0.0
            )
            currency = settlement.currency_id
            has_actuals = bool(settlement.statement_order_count) or not (
                currency.is_zero(settlement.actual_customer_collections)
                and currency.is_zero(settlement.actual_contribution)
                and currency.is_zero(settlement.actual_commission)
            ) if currency else False
            settlement.customer_collections_mismatch = has_actuals and not currency.is_zero(
                settlement.actual_customer_collections - settlement.expected_customer_collections
            )
            settlement.contribution_mismatch = has_actuals and not currency.is_zero(
                settlement.actual_contribution - settlement.expected_contribution
            )
            settlement.commission_mismatch = has_actuals and not currency.is_zero(
                settlement.actual_commission - settlement.expected_commission
            )

    @api.constrains("date_start", "date_end")
    def _check_dates(self):
        for settlement in self:
            if settlement.date_start > settlement.date_end:
                raise ValidationError(_("Settlement start date cannot be after its end date."))

    @api.constrains(
        "actual_customer_collections", "actual_contribution", "actual_commission"
    )
    def _check_actual_amounts(self):
        for settlement in self:
            if any(amount < 0 for amount in (
                settlement.actual_customer_collections,
                settlement.actual_contribution,
                settlement.actual_commission,
            )):
                raise ValidationError(_("Actual settlement amounts cannot be negative."))

    def _get_campaign_lines(self):
        self.ensure_one()
        timezone = pytz.timezone(self.env.user.tz or "UTC")
        start = timezone.localize(datetime.combine(self.date_start, time.min)).astimezone(
            pytz.UTC
        ).replace(tzinfo=None)
        end = timezone.localize(datetime.combine(self.date_end, time.max)).astimezone(
            pytz.UTC
        ).replace(tzinfo=None)
        return self.env["pos.order.line"].search([
            ("online_aggregator_id", "=", self.aggregator_id.id),
            ("order_id.company_id", "=", self.company_id.id),
            ("order_id.currency_id", "=", self.currency_id.id),
            ("order_id.date_order", ">=", start),
            ("order_id.date_order", "<=", end),
            ("order_id.state", "in", ("paid", "done")),
        ])

    def action_load_expected(self):
        for settlement in self:
            if settlement.state == "reconciled":
                raise UserError(_("A reconciled settlement cannot be recalculated."))

            report_lines = self.env["online.campaign.performance.report"].search([
                ("aggregator_id", "=", settlement.aggregator_id.id),
                ("company_id", "=", settlement.company_id.id),
                ("currency_id", "=", settlement.currency_id.id),
                ("date", ">=", settlement.date_start),
                ("date", "<=", settlement.date_end),
            ])

            settlement.write({
                "order_count": int(sum(report_lines.mapped("order_count"))),
                "line_count": int(sum(report_lines.mapped("line_count"))),
                "expected_customer_collections": settlement.currency_id.round(sum(
                    report_lines.mapped("customer_collections")
                )),
                "expected_contribution": settlement.currency_id.round(sum(
                    report_lines.mapped("aggregator_contribution")
                )),
                "expected_commission": settlement.currency_id.round(sum(
                    report_lines.mapped("estimated_commission")
                )),
            })

    def _parse_statement_rows(self):
        """Read the uploaded statement and return its header index and data rows."""
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_(
                "The openpyxl Python library is required to import statement files."
            ))
        if not self.statement_file:
            raise UserError(_("Upload a statement file first."))
        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(base64.b64decode(self.statement_file)),
                read_only=True, data_only=True,
            )
        except Exception:
            raise UserError(_("The statement file could not be read as an Excel file."))
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise UserError(_("The statement file is empty."))
        header_index = {}
        for position, header in enumerate(header_row):
            if header:
                header_index[str(header)] = position
        missing = [
            label for label in STATEMENT_COLUMNS.values() if label not in header_index
        ]
        if missing:
            raise UserError(_(
                "This file doesn't match the expected Talabat statement format. "
                "Missing columns: %s", ", ".join(missing)
            ))
        return header_index, rows

    def _row_in_date_range(self, row, header_index):
        """True if the statement row's Date / Time falls within date_start/date_end."""
        self.ensure_one()
        value = row[header_index[STATEMENT_COLUMNS["date"]]]
        if isinstance(value, datetime):
            row_date = value.date()
        elif isinstance(value, str) and value:
            try:
                row_date = datetime.fromisoformat(value).date()
            except ValueError:
                return False
        else:
            return False
        return self.date_start <= row_date <= self.date_end

    def action_download_statement_template(self):
        """Generate a blank statement template with the columns the importer reads.

        Lets any aggregator's statement (Careem, MyThings, etc.) be reshaped into
        this layout by hand before it's uploaded and imported.
        """
        if openpyxl is None:
            raise UserError(_(
                "The openpyxl Python library is required to generate the template."
            ))
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Statement"
        headers = list(STATEMENT_COLUMNS.values())
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        sheet.append([
            "3665735758", datetime(2026, 6, 1, 16, 55, 56), "Successful",
            28.51, 0.0, 0.0, 0.0, 0.387, 0.06192, 0.0, 4.57, 0.7312,
        ])
        sheet.append([
            "3666117400", datetime(2026, 6, 1, 19, 36, 57), STATEMENT_CANCELLED_STATUS,
            0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.924, 0.14784,
        ])
        for column_cells in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = max(width + 2, 12)

        buffer = io.BytesIO()
        workbook.save(buffer)
        attachment = self.env["ir.attachment"].create({
            "name": "Aggregator Statement Template.xlsx",
            "type": "binary",
            "datas": base64.b64encode(buffer.getvalue()),
            "mimetype": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        })
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def action_import_statement(self):
        """Read the uploaded aggregator statement and fill in the Actual fields."""
        for settlement in self:
            if settlement.state == "reconciled":
                raise UserError(_("A reconciled settlement is locked."))
            header_index, rows = settlement._parse_statement_rows()
            currency = settlement.currency_id

            totals = {key: 0.0 for key in STATEMENT_COLUMNS if key not in ("order_id", "date", "status")}
            order_count = 0
            skipped_out_of_range = 0
            cancelled_count = 0
            cancelled_commission = 0.0

            def value_at(row, key):
                return row[header_index[STATEMENT_COLUMNS[key]]] or 0.0

            for row in rows:
                order_id = row[header_index[STATEMENT_COLUMNS["order_id"]]]
                if not order_id:
                    continue
                if not settlement._row_in_date_range(row, header_index):
                    skipped_out_of_range += 1
                    continue
                order_count += 1
                status = row[header_index[STATEMENT_COLUMNS["status"]]]
                commission_row = value_at(row, "commission") + value_at(row, "commission_vat")

                for key in totals:
                    totals[key] += value_at(row, key)

                if status == STATEMENT_CANCELLED_STATUS:
                    cancelled_count += 1
                    cancelled_commission += commission_row

            settlement.write({
                "statement_order_count": order_count,
                "actual_customer_collections": currency.round(totals["total_sale"]),
                "actual_contribution": currency.round(totals["voucher_cost_to_talabat"]),
                "actual_commission": currency.round(
                    totals["commission"] + totals["commission_vat"]
                ),
                "actual_tpro_loyalty_charges": currency.round(totals["loyalty_charges"]),
                "actual_tpro_cost_covered": currency.round(totals["tpro_cost_covered"]),
                "actual_sponsored_deal_fees": currency.round(totals["sponsored_deal_fees"]),
                "actual_payment_handling_charges": currency.round(
                    totals["payment_handling_charges"]
                ),
                "actual_payment_handling_charges_vat": currency.round(
                    totals["payment_handling_charges_vat"]
                ),
                "cancelled_charged_order_count": cancelled_count,
                "cancelled_charged_commission": currency.round(cancelled_commission),
            })
            if skipped_out_of_range:
                settlement.message_post(body=_(
                    "Statement import: %(kept)s rows within %(start)s - %(end)s used, "
                    "%(skipped)s rows outside that range excluded.",
                    kept=order_count, start=settlement.date_start, end=settlement.date_end,
                    skipped=skipped_out_of_range,
                ))

    def _get_pos_order_totals_by_id(self):
        """Per-order customer collections / contribution / commission for this
        settlement's aggregator and date range, keyed by pos.order id.

        Mirrors the campaign and non-campaign formulas used by
        online.campaign.performance.report, but grouped per order instead of
        per date/campaign/session, so it can be matched 1:1 against the
        aggregator's statement rows.
        """
        self.ensure_one()
        timezone = pytz.timezone(self.env.user.tz or "UTC")
        start = timezone.localize(datetime.combine(self.date_start, time.min)).astimezone(
            pytz.UTC
        ).replace(tzinfo=None)
        end = timezone.localize(datetime.combine(self.date_end, time.max)).astimezone(
            pytz.UTC
        ).replace(tzinfo=None)

        self.env.cr.execute("""
            WITH campaign_rows AS (
                SELECT
                    po.id AS order_id,
                    SUM(pol.online_customer_paid_amount) AS customer_collections,
                    SUM(COALESCE(pol.aggregator_contribution_amount, 0.0)) AS contribution,
                    SUM(COALESCE(pol.aggregator_commission_amount, 0.0)) AS commission
                FROM pos_order_line pol
                JOIN pos_order po ON po.id = pol.order_id
                WHERE pol.online_aggregator_id = %(aggregator_id)s
                  AND po.company_id = %(company_id)s
                  AND po.date_order >= %(start)s AND po.date_order <= %(end)s
                  AND po.state IN ('paid', 'done')
                GROUP BY po.id
            ),
            non_campaign_rows AS (
                SELECT
                    po.id AS order_id,
                    pay.paid_amount AS customer_collections,
                    0.0 AS contribution,
                    (
                        CASE
                            WHEN ag.commission_base = 'before_tax' THEN
                                CASE
                                    WHEN ABS(po.amount_total) = 0 THEN 0.0
                                    ELSE ABS(pay.paid_amount) * ABS(po.amount_total - po.amount_tax) / ABS(po.amount_total)
                                END
                            ELSE ABS(pay.paid_amount)
                        END
                    ) * ag.default_commission_percent / 100.0 AS commission
                FROM (
                    SELECT
                        pp.pos_order_id AS order_id,
                        SUM(pp.amount) AS paid_amount
                    FROM pos_payment pp
                    JOIN online_campaign_aggregator_pos_payment_method_rel rel
                         ON rel.payment_method_id = pp.payment_method_id
                    WHERE rel.aggregator_id = %(aggregator_id)s
                    GROUP BY pp.pos_order_id
                ) pay
                JOIN pos_order po ON po.id = pay.order_id
                JOIN online_campaign_aggregator ag ON ag.id = %(aggregator_id)s
                WHERE po.company_id = %(company_id)s
                  AND po.date_order >= %(start)s AND po.date_order <= %(end)s
                  AND po.state IN ('paid', 'done')
                  AND NOT EXISTS (
                      SELECT 1 FROM pos_order_line campaign_line
                       WHERE campaign_line.order_id = po.id
                         AND campaign_line.online_campaign_id IS NOT NULL
                  )
            )
            SELECT order_id, SUM(customer_collections), SUM(contribution), SUM(commission)
              FROM (SELECT * FROM campaign_rows UNION ALL SELECT * FROM non_campaign_rows) all_rows
            GROUP BY order_id
        """, {
            "aggregator_id": self.aggregator_id.id,
            "company_id": self.company_id.id,
            "start": start,
            "end": end,
        })
        totals_by_order_id = {
            row[0]: {"customer_collections": row[1] or 0.0, "contribution": row[2] or 0.0, "commission": row[3] or 0.0}
            for row in self.env.cr.fetchall()
        }
        if not totals_by_order_id:
            return {}

        orders = self.env["pos.order"].browse(totals_by_order_id.keys())
        return {
            order.id: {
                "name": order.name,
                "reference": (order.online_aggregator_order_ref or "").strip(),
                **totals_by_order_id[order.id],
            }
            for order in orders
        }

    def action_export_order_comparison(self):
        """Export a 3-sheet order-by-order reconciliation: Matched, Talabat
        rows without a clean POS match, and POS orders without a clean
        Talabat match. Orders match on the aggregator's own order reference
        (online_aggregator_order_ref), entered in the back office per order.
        """
        self.ensure_one()
        if openpyxl is None:
            raise UserError(_(
                "The openpyxl Python library is required to export the order comparison."
            ))
        header_index, rows = self._parse_statement_rows()
        currency = self.currency_id

        def value_at(row, key):
            return row[header_index[STATEMENT_COLUMNS[key]]] or 0.0

        talabat_by_ref = {}
        for row in rows:
            order_ref = row[header_index[STATEMENT_COLUMNS["order_id"]]]
            if not order_ref:
                continue
            if not self._row_in_date_range(row, header_index):
                continue
            order_ref = str(order_ref)
            entry = talabat_by_ref.setdefault(order_ref, {
                "status": row[header_index[STATEMENT_COLUMNS["status"]]],
                "total_sale": 0.0, "contribution": 0.0, "commission": 0.0,
            })
            entry["total_sale"] += value_at(row, "total_sale")
            entry["contribution"] += value_at(row, "voucher_cost_to_talabat")
            entry["commission"] += value_at(row, "commission") + value_at(row, "commission_vat")

        pos_totals = self._get_pos_order_totals_by_id()
        pos_by_ref = {}
        pos_no_ref = []
        for data in pos_totals.values():
            if data["reference"]:
                pos_by_ref.setdefault(data["reference"], []).append(data)
            else:
                pos_no_ref.append(data)

        def is_match(talabat_entry, pos_entry):
            return (
                currency.is_zero(talabat_entry["total_sale"] - pos_entry["customer_collections"])
                and currency.is_zero(talabat_entry["contribution"] - pos_entry["contribution"])
                and currency.is_zero(talabat_entry["commission"] - pos_entry["commission"])
            )

        matched_refs = set()
        for ref, talabat_entry in talabat_by_ref.items():
            candidates = pos_by_ref.get(ref) or []
            if len(candidates) == 1 and is_match(talabat_entry, candidates[0]):
                matched_refs.add(ref)

        workbook = openpyxl.Workbook()
        bold = openpyxl.styles.Font(bold=True)

        def write_sheet(sheet, headers, data_rows):
            sheet.append(headers)
            for cell in sheet[1]:
                cell.font = bold
            for data_row in data_rows:
                sheet.append(data_row)
            for column_cells in sheet.columns:
                width = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column_letter].width = max(width + 2, 10)

        matched_sheet = workbook.active
        matched_sheet.title = "Matched"
        matched_headers = [
            "Aggregator Order Ref", "POS Order", "Total Sale / Customer Collections",
            "Contribution", "Commission",
        ]
        matched_rows = []
        for ref in sorted(matched_refs):
            talabat_entry = talabat_by_ref[ref]
            pos_entry = pos_by_ref[ref][0]
            matched_rows.append([
                ref, pos_entry["name"],
                currency.round(talabat_entry["total_sale"]),
                currency.round(talabat_entry["contribution"]),
                currency.round(talabat_entry["commission"]),
            ])
        write_sheet(matched_sheet, matched_headers, matched_rows)

        talabat_sheet = workbook.create_sheet("Talabat Not Matching POS")
        talabat_headers = [
            "Aggregator Order Ref", "Talabat Status", "Talabat Total Sale", "POS Customer Collections",
            "Talabat Contribution", "POS Contribution", "Talabat Commission", "POS Commission",
            "POS Order", "Issue",
        ]
        talabat_rows = []
        for ref, talabat_entry in talabat_by_ref.items():
            if ref in matched_refs:
                continue
            candidates = pos_by_ref.get(ref) or []
            if not candidates:
                pos_entry, issue = None, "Not found in POS"
            elif len(candidates) > 1:
                pos_entry, issue = candidates[0], "Multiple POS orders share this reference"
            else:
                pos_entry, issue = candidates[0], "Amount mismatch"
            talabat_rows.append([
                ref, talabat_entry["status"],
                currency.round(talabat_entry["total_sale"]),
                currency.round(pos_entry["customer_collections"]) if pos_entry else None,
                currency.round(talabat_entry["contribution"]),
                currency.round(pos_entry["contribution"]) if pos_entry else None,
                currency.round(talabat_entry["commission"]),
                currency.round(pos_entry["commission"]) if pos_entry else None,
                pos_entry["name"] if pos_entry else "",
                issue,
            ])
        write_sheet(talabat_sheet, talabat_headers, talabat_rows)

        pos_sheet = workbook.create_sheet("POS Not Matching Talabat")
        pos_headers = [
            "POS Order", "Aggregator Order Ref", "POS Customer Collections", "Talabat Total Sale",
            "POS Contribution", "Talabat Contribution", "POS Commission", "Talabat Commission", "Issue",
        ]
        pos_rows = []
        for ref, candidates in pos_by_ref.items():
            if ref in matched_refs:
                continue
            talabat_entry = talabat_by_ref.get(ref)
            for pos_entry in candidates:
                issue = "Not found in Talabat statement" if not talabat_entry else "Amount mismatch"
                pos_rows.append([
                    pos_entry["name"], ref,
                    currency.round(pos_entry["customer_collections"]),
                    currency.round(talabat_entry["total_sale"]) if talabat_entry else None,
                    currency.round(pos_entry["contribution"]),
                    currency.round(talabat_entry["contribution"]) if talabat_entry else None,
                    currency.round(pos_entry["commission"]),
                    currency.round(talabat_entry["commission"]) if talabat_entry else None,
                    issue,
                ])
        for pos_entry in pos_no_ref:
            pos_rows.append([
                pos_entry["name"], "",
                currency.round(pos_entry["customer_collections"]), None,
                currency.round(pos_entry["contribution"]), None,
                currency.round(pos_entry["commission"]), None,
                "No aggregator order reference entered on this POS order",
            ])
        write_sheet(pos_sheet, pos_headers, pos_rows)

        buffer = io.BytesIO()
        workbook.save(buffer)
        attachment = self.env["ir.attachment"].create({
            "name": f"{self.name} Order Comparison.xlsx",
            "type": "binary",
            "datas": base64.b64encode(buffer.getvalue()),
            "mimetype": (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        })
        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/{attachment.id}?download=true",
            "target": "self",
        }

    def action_confirm(self):
        for settlement in self:
            overlap = self.search_count([
                ("id", "!=", settlement.id),
                ("aggregator_id", "=", settlement.aggregator_id.id),
                ("company_id", "=", settlement.company_id.id),
                ("currency_id", "=", settlement.currency_id.id),
                ("state", "in", ("confirmed", "reconciled")),
                ("date_start", "<=", settlement.date_end),
                ("date_end", ">=", settlement.date_start),
            ])
            if overlap:
                raise UserError(_(
                    "This period overlaps another confirmed or reconciled settlement for the aggregator."
                ))
        self.action_load_expected()
        self.write({"state": "confirmed"})

    def action_mark_reconciled(self):
        if not self.env.user.has_group(
            "online_campaigns_discount.group_online_campaign_finance_manager"
        ):
            raise AccessError(_("Only an Online Campaign Finance Manager can reconcile settlements."))
        for settlement in self:
            if settlement.state != "confirmed":
                raise UserError(_("Confirm the settlement before reconciliation."))
            if not settlement.statement_reference:
                raise UserError(_("Enter the aggregator statement reference."))
            if not settlement.bank_statement_line_id and not settlement.account_move_id:
                raise UserError(_(
                    "Link the received bank statement line or the settlement/commission accounting entry."
                ))
            has_component_mismatch = (
                settlement.customer_collections_mismatch
                or settlement.contribution_mismatch
                or settlement.commission_mismatch
            )
            if (
                not settlement.currency_id.is_zero(settlement.variance_amount)
                or has_component_mismatch
            ) and not settlement.variance_reason:
                raise UserError(_("Explain the settlement variance before reconciliation."))
            settlement.write({
                "state": "reconciled", "reconciled_by": self.env.user.id,
                "reconciled_at": fields.Datetime.now(),
            })

    def action_reset_to_draft(self):
        if not self.env.user.has_group(
            "online_campaigns_discount.group_online_campaign_finance_manager"
        ):
            raise AccessError(_("Only an Online Campaign Finance Manager can reset settlements."))
        self.filtered(lambda settlement: settlement.state != "reconciled").write({"state": "draft"})
