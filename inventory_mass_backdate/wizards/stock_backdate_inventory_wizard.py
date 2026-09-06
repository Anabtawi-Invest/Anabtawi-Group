import base64
import io
import logging
from datetime import datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools import float_compare, float_is_zero

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None


class StockBackdateInventoryWizard(models.TransientModel):
    _name = 'stock.backdate.inventory.wizard'
    _description = 'Apply Backdated Physical Count'

    count_date = fields.Datetime(
        string='Count Date',
        required=True,
        default=fields.Datetime.now,
        help="Date the physical count was actually taken. All adjustments, "
             "stock moves and accounting entries created by this wizard are "
             "dated to this date. Counted quantities are compared against the "
             "on-hand quantity as it stood on this date, so any movements "
             "between the count date and today are preserved (not overwritten). "
             "A per-row 'Count Date' in the Excel file overrides this.",
    )
    reason = fields.Char(
        string='Reason',
        required=True,
        default='Backdated physical inventory count',
    )
    source = fields.Selection(
        [('selection', 'Selected lines on screen'),
         ('excel', 'Upload an Excel file')],
        string='Counts From',
        required=True,
        default='selection',
    )
    import_file = fields.Binary(string='Excel File')
    import_filename = fields.Char(string='File Name')
    location_id = fields.Many2one(
        'stock.location',
        string='Default Location',
        domain="[('usage', '=', 'internal')]",
        help="Location used for rows in the Excel file that do not name one.",
    )
    line_count = fields.Integer(
        string='Counted Lines',
        compute='_compute_preview',
    )
    skipped_no_count = fields.Integer(
        string='Lines Without a Count',
        compute='_compute_preview',
    )

    # ------------------------------------------------------------------
    # Preview
    # ------------------------------------------------------------------
    def _get_target_quants(self):
        """Quants selected on the Physical Inventory screen."""
        if self.env.context.get('active_model') != 'stock.quant':
            return self.env['stock.quant']
        quant_ids = self.env.context.get('active_ids') or []
        return self.env['stock.quant'].browse(quant_ids).exists()

    @api.depends('source', 'count_date')
    def _compute_preview(self):
        quants = self._get_target_quants()
        for wizard in self:
            if wizard.source == 'selection':
                counted = quants.filtered('inventory_quantity_set')
                wizard.line_count = len(counted)
                wizard.skipped_no_count = len(quants) - len(counted)
            else:
                wizard.line_count = 0
                wizard.skipped_no_count = 0

    # ------------------------------------------------------------------
    # Core: apply one count (product/location/lot) gap-aware
    # ------------------------------------------------------------------
    def _historical_on_hand(self, quant, count_date):
        """On-hand qty for this product/location/lot as of ``count_date``."""
        product = quant.product_id
        ctx = {'to_date': count_date, 'location': quant.location_id.id}
        if quant.lot_id:
            ctx['lot_id'] = quant.lot_id.id
        if quant.owner_id:
            ctx['owner_id'] = quant.owner_id.id
        if quant.package_id:
            ctx['package_id'] = quant.package_id.id
        return product.with_context(**ctx).qty_available

    def _svls_of(self, move):
        """Valuation layers of a move (no reverse field on stock.move)."""
        return self.env['stock.valuation.layer'].sudo().search(
            [('stock_move_id', '=', move.id)])

    def _backdate_valuation(self, move, count_date):
        """Push valuation layers and their journal entries to ``count_date``.

        Odoo stamps a valuation layer with its ``create_date`` and the linked
        accounting entry with 'today'. To land the count in the right period
        we rewrite both. ``create_date`` is a magic column, so it can only be
        set with raw SQL.
        """
        svls = self._svls_of(move)
        if not svls:
            return
        self.env.cr.execute(
            "UPDATE stock_valuation_layer SET create_date = %s WHERE id IN %s",
            (count_date, tuple(svls.ids)),
        )
        for entry in svls.mapped('account_move_id').filtered(lambda m: m):
            if any(line.reconciled for line in entry.line_ids):
                _logger.info(
                    "Backdate inventory: journal entry %s has reconciled "
                    "lines, leaving its date untouched.", entry.name)
                continue
            try:
                entry.write({'date': count_date})
            except UserError as err:
                raise UserError(_(
                    "Cannot date the accounting entry for %(product)s to "
                    "%(date)s. The accounting period is probably locked.\n\n"
                    "Original error: %(err)s",
                    product=move.product_id.display_name,
                    date=count_date, err=err,
                )) from err

    def _create_backdated_move(self, quant, delta, count_date):
        product = quant.product_id
        company = quant.company_id or self.env.company
        inventory_loc = product.with_company(company).property_stock_inventory
        if float_compare(delta, 0.0, precision_rounding=product.uom_id.rounding) > 0:
            loc_src, loc_dst = inventory_loc, quant.location_id
        else:
            loc_src, loc_dst = quant.location_id, inventory_loc

        move_vals = quant._get_inventory_move_values(abs(delta), loc_src, loc_dst)
        move_vals.update({
            'date': count_date,
            'origin': _('Backdated count: %s', self.reason),
        })
        move = self.env['stock.move'].with_context(
            inventory_mode=True).create(move_vals)
        move._action_done()
        move.write({'date': count_date})
        move.move_line_ids.write({'date': count_date})
        self._backdate_valuation(move, count_date)
        return move

    def _apply_one(self, quant, counted, count_date):
        """Apply one gap-aware backdated count. Returns an audit dict."""
        product = quant.product_id
        rounding = product.uom_id.rounding
        hist_qty = self._historical_on_hand(quant, count_date)
        delta = counted - hist_qty

        audit = {
            'default_code': product.default_code or '',
            'product': product.display_name,
            'location': quant.location_id.complete_name,
            'lot': quant.lot_id.name or '',
            'count_date': count_date,
            'hist_qty': hist_qty,
            'counted': counted,
            'delta': delta,
            'uom': product.uom_id.name,
            'value': 0.0,
            'move': '',
            'entry': '',
        }
        if float_is_zero(delta, precision_rounding=rounding):
            audit['note'] = 'Already matched (no change)'
            return audit

        move = self._create_backdated_move(quant, delta, count_date)
        svls = self._svls_of(move)
        audit['value'] = sum(svls.mapped('value'))
        audit['move'] = move.reference or move.name or str(move.id)
        entries = svls.mapped('account_move_id')
        audit['entry'] = ', '.join(entries.mapped('name')) if entries else ''
        audit['note'] = 'Adjusted'
        return audit

    # ------------------------------------------------------------------
    # Build lines from the two sources
    # ------------------------------------------------------------------
    def _lines_from_selection(self):
        quants = self._get_target_quants().filtered('inventory_quantity_set')
        if not quants:
            raise UserError(_(
                "None of the selected lines have a counted quantity. Enter a "
                "value in the 'Counted' column first, then run this again."))
        lines = []
        for quant in quants:
            count_date = quant.counted_date or self.count_date
            lines.append((quant, quant.inventory_quantity, count_date, quant))
        return lines

    def _find_product(self, ref):
        Product = self.env['product.product']
        ref = str(ref).strip()
        product = Product.search([('default_code', '=', ref)], limit=1)
        if not product:
            product = Product.search([('barcode', '=', ref)], limit=1)
        if not product:
            product = Product.search([('name', '=', ref)], limit=1)
        if not product:
            raise UserError(_(
                "No product found for '%s'. Match is by Internal Reference, "
                "then Barcode, then exact Name.", ref))
        return product

    def _find_location(self, name):
        if not name:
            if not self.location_id:
                raise UserError(_(
                    "A row has no location and no Default Location is set on "
                    "the wizard."))
            return self.location_id
        Location = self.env['stock.location']
        name = str(name).strip()
        loc = Location.search([('complete_name', '=', name)], limit=1)
        if not loc:
            loc = Location.search([('name', '=', name)], limit=1)
        if not loc:
            raise UserError(_("No internal location found named '%s'.", name))
        return loc

    def _parse_date(self, value):
        if not value:
            return self.count_date
        if isinstance(value, datetime):
            return value
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
            try:
                return datetime.strptime(str(value).strip(), fmt)
            except ValueError:
                continue
        raise UserError(_(
            "Could not read the count date '%s'. Use YYYY-MM-DD.", value))

    def _lines_from_excel(self):
        if openpyxl is None:
            raise UserError(_(
                "The Python library 'openpyxl' is not available on the server, "
                "so Excel upload cannot be used."))
        if not self.import_file:
            raise UserError(_("Attach an Excel file first."))
        data = base64.b64decode(self.import_file)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        except Exception as err:
            raise UserError(_("Could not read the Excel file: %s", err))
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 2:
            raise UserError(_("The Excel file has no data rows."))

        header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

        def col(*names):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        i_ref = col('internal reference', 'reference', 'default_code', 'code', 'product code')
        i_qty = col('counted quantity', 'counted', 'quantity', 'qty', 'count')
        i_loc = col('location')
        i_lot = col('lot/serial', 'lot', 'serial', 'lot/serial number')
        i_date = col('count date', 'counted on', 'date')
        if i_ref is None or i_qty is None:
            raise UserError(_(
                "The Excel file must have at least an 'Internal Reference' "
                "column and a 'Counted Quantity' column. Download the template "
                "for the exact layout."))

        lines = []
        for n, row in enumerate(rows[1:], start=2):
            if row is None or all(c is None for c in row):
                continue
            ref = row[i_ref]
            qty = row[i_qty]
            if ref is None or qty is None or str(ref).strip() == '':
                continue
            try:
                counted = float(qty)
            except (TypeError, ValueError):
                raise UserError(_(
                    "Row %(row)s: '%(val)s' is not a valid counted quantity.",
                    row=n, val=qty))
            product = self._find_product(ref)
            location = self._find_location(row[i_loc] if i_loc is not None else None)
            lot = self.env['stock.lot']
            if i_lot is not None and row[i_lot]:
                lot = self.env['stock.lot'].search([
                    ('name', '=', str(row[i_lot]).strip()),
                    ('product_id', '=', product.id),
                ], limit=1)
                if not lot:
                    raise UserError(_(
                        "Row %(row)s: lot/serial '%(lot)s' not found for %(p)s.",
                        row=n, lot=row[i_lot], p=product.display_name))
            count_date = self._parse_date(row[i_date] if i_date is not None else None)

            quant = self._get_or_create_quant(product, location, lot)
            lines.append((quant, counted, count_date, None))
        if not lines:
            raise UserError(_("No usable count rows found in the Excel file."))
        return lines

    def _get_or_create_quant(self, product, location, lot):
        Quant = self.env['stock.quant']
        domain = [
            ('product_id', '=', product.id),
            ('location_id', '=', location.id),
            ('lot_id', '=', lot.id if lot else False),
        ]
        quant = Quant.search(domain, limit=1)
        if not quant:
            quant = Quant.with_context(inventory_mode=True).create({
                'product_id': product.id,
                'location_id': location.id,
                'lot_id': lot.id if lot else False,
            })
        return quant

    # ------------------------------------------------------------------
    # Main action
    # ------------------------------------------------------------------
    def action_apply(self):
        self.ensure_one()
        if self.source == 'excel':
            lines = self._lines_from_excel()
        else:
            lines = self._lines_from_selection()

        audit_rows, applied, zero_delta = [], 0, 0
        for quant, counted, count_date, screen_quant in lines:
            if not count_date:
                raise UserError(_(
                    "No count date for %s.", quant.product_id.display_name))
            audit = self._apply_one(quant, counted, count_date)
            audit_rows.append(audit)
            if audit.get('note') == 'Adjusted':
                applied += 1
            else:
                zero_delta += 1
            if screen_quant is not None:
                screen_quant.write({
                    'counted_date': count_date,
                    'inventory_quantity': 0,
                    'inventory_quantity_set': False,
                })

        attachment = self._build_audit_file(audit_rows)
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }

    def action_load_to_screen(self):
        """Excel -> populate the 'Counted' column on Physical Inventory.

        Does NOT post anything. The user reviews on screen, then runs
        'Apply Backdated Count' from the list to post + get the audit file.
        """
        self.ensure_one()
        lines = self._lines_from_excel()
        quant_ids = []
        for quant, counted, count_date, _screen in lines:
            quant.with_context(inventory_mode=True).write({
                'inventory_quantity': counted,
                'counted_date': count_date,
            })
            quant_ids.append(quant.id)
        return {
            'type': 'ir.actions.act_window',
            'name': _('Physical Inventory'),
            'res_model': 'stock.quant',
            'view_mode': 'list',
            'domain': [('id', 'in', quant_ids)],
            'context': {'inventory_mode': True},
            'views': [(
                self.env.ref(
                    'stock.view_stock_quant_tree_inventory_editable').id,
                'list')],
            'help': _("Counts loaded from Excel. Review them, then select all "
                      "and run Action ▸ Apply Backdated Count."),
        }

    # ------------------------------------------------------------------
    # Excel audit output
    # ------------------------------------------------------------------
    def _build_audit_file(self, audit_rows):
        if xlsxwriter is None:
            raise UserError(_(
                "The Python library 'xlsxwriter' is not available on the "
                "server, so the audit Excel file cannot be produced."))
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        ws = wb.add_worksheet('Backdate Audit')
        bold = wb.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1})
        cell = wb.add_format({'border': 1})
        num = wb.add_format({'border': 1, 'num_format': '#,##0.000'})
        money = wb.add_format({'border': 1, 'num_format': '#,##0.00'})
        datef = wb.add_format({'border': 1, 'num_format': 'yyyy-mm-dd hh:mm'})

        headers = [
            'Internal Reference', 'Product', 'Location', 'Lot/Serial',
            'Count Date', 'On-hand @ Count Date', 'Counted', 'Adjustment',
            'UoM', 'Value Change', 'Stock Move', 'Journal Entry', 'Result',
        ]
        widths = [18, 40, 24, 14, 18, 18, 12, 12, 8, 14, 16, 16, 22]
        for c, (h, w) in enumerate(zip(headers, widths)):
            ws.write(0, c, h, bold)
            ws.set_column(c, c, w)
        ws.freeze_panes(1, 0)

        for r, a in enumerate(audit_rows, start=1):
            cd = a['count_date']
            if isinstance(cd, str):
                cd = self._parse_date(cd)
            ws.write(r, 0, a['default_code'], cell)
            ws.write(r, 1, a['product'], cell)
            ws.write(r, 2, a['location'], cell)
            ws.write(r, 3, a['lot'], cell)
            ws.write_datetime(r, 4, cd, datef)
            ws.write_number(r, 5, a['hist_qty'], num)
            ws.write_number(r, 6, a['counted'], num)
            ws.write_number(r, 7, a['delta'], num)
            ws.write(r, 8, a['uom'], cell)
            ws.write_number(r, 9, a['value'], money)
            ws.write(r, 10, a['move'], cell)
            ws.write(r, 11, a['entry'], cell)
            ws.write(r, 12, a.get('note', ''), cell)

        footer = len(audit_rows) + 2
        ws.write(footer, 0, 'Generated by', bold)
        ws.write(footer, 1, self.env.user.name, cell)
        ws.write(footer + 1, 0, 'Generated on', bold)
        ws.write_datetime(footer + 1, 1, fields.Datetime.now(), datef)
        ws.write(footer + 2, 0, 'Reason', bold)
        ws.write(footer + 2, 1, self.reason or '', cell)
        wb.close()
        output.seek(0)

        fname = 'backdate_audit_%s.xlsx' % fields.Datetime.now().strftime(
            '%Y%m%d_%H%M%S')
        return self.env['ir.attachment'].create({
            'name': fname,
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sheet',
        })

    # ------------------------------------------------------------------
    # Template download
    # ------------------------------------------------------------------
    def action_download_template(self):
        self.ensure_one()
        if xlsxwriter is None:
            raise UserError(_("xlsxwriter is not available on the server."))
        output = io.BytesIO()
        wb = xlsxwriter.Workbook(output, {'in_memory': True})
        ws = wb.add_worksheet('Counts')
        bold = wb.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1})
        cell = wb.add_format({'border': 1})
        headers = ['Internal Reference', 'Product', 'Location',
                   'Lot/Serial', 'Counted Quantity', 'Count Date']
        widths = [18, 40, 24, 14, 16, 16]
        for c, (h, w) in enumerate(zip(headers, widths)):
            ws.write(0, c, h, bold)
            ws.set_column(c, c, w)
        example = ['1010085', 'سكر', 'السلط 1/Stock', '', 634.5, '2025-08-31']
        for c, v in enumerate(example):
            ws.write(1, c, v, cell)
        ws.write(3, 0, 'Only "Internal Reference" and "Counted Quantity" are '
                       'required. Location falls back to the wizard default. '
                       'Count Date falls back to the wizard Count Date.', cell)
        wb.close()
        output.seek(0)
        attachment = self.env['ir.attachment'].create({
            'name': 'backdate_count_template.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(output.read()),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.'
                        'spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }
